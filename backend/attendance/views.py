from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from datetime import date, timedelta, datetime



User = get_user_model()

from .permissions import HasAttendanceViewPermission
from .models import Attendance, StudentAttendance, Weekend, ZKTecoDevice, ZKTecoEmployeeMapping, StaffAttendance, EmployeeShiftTiming, Holiday
from .serializers import (
    AttendanceSerializer,
    StudentAttendanceSerializer,
    AttendanceMarkingSerializer,
    AttendanceSummarySerializer,
    ZKTecoDeviceSerializer,
    ZKTecoMappingSerializer,
    StaffAttendanceSerializer,
)
from students.models import Student
from classes.models import ClassRoom
from teachers.models import Teacher
from coordinator.models import Coordinator
from notifications.services import create_notification
from .services.alerts import process_consecutive_absence_alerts
from .services import calendar_utils
from .services.holiday_utils import (
    collect_shifts_from_levels,
    normalize_shift_value,  
    resolve_allowed_shifts,
    validate_grades_for_levels,
    validate_levels_for_shift,
)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_attendance(request):
    """
    Mark attendance for a class on a specific date
    """
    try:
        serializer = AttendanceMarkingSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        classroom_id = data['classroom_id']
        date = data['date']
        student_attendance_data = data['student_attendance']
        
        classroom = get_object_or_404(ClassRoom, id=classroom_id)
        
        # Check if date is a holiday (support multiple levels and grade-specific)
        from .models import Holiday
        level = classroom.grade.level if classroom.grade else None
        grade = classroom.grade if classroom.grade else None
        
        if level:
            # Check for holidays matching this level and date
            holidays = Holiday.objects.filter(
                date=date
            ).filter(
                Q(levels=level) | Q(level=level)  # Support both old and new fields
            ).distinct()
            
            # Check if any holiday applies to this classroom
            for holiday in holidays:
                # If holiday has specific grades, check if this classroom's grade is included
                if holiday.grades.exists():
                    if grade and grade in holiday.grades.all():
                        return Response({
                            'error': f'This date is a holiday: {holiday.reason}. Attendance marking is disabled.',
                            'is_holiday': True,
                            'holiday_reason': holiday.reason
                        }, status=status.HTTP_400_BAD_REQUEST)
                else:
                    # No specific grades - applies to all grades in the level
                    return Response({
                        'error': f'This date is a holiday: {holiday.reason}. Attendance marking is disabled.',
                        'is_holiday': True,
                        'holiday_reason': holiday.reason
                    }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get teacher from request user
        try:
            # Find teacher by employee code (username) since there's no direct relationship
            from teachers.models import Teacher
            teacher = Teacher.objects.get(employee_code=request.user.username)
        except Teacher.DoesNotExist:
            teacher = None
        
        with transaction.atomic():
            # Create or get attendance record
            # marked_by must be a User instance, not Teacher
            user = request.user
            if teacher and hasattr(teacher, 'user') and teacher.user is not None:
                marked_by_user = teacher.user
            else:
                marked_by_user = user
                
            # Automatically set organization if user belongs to one
            org = None
            if not user.is_superadmin() and user.organization:
                org = user.organization
            
            attendance, created = Attendance.objects.get_or_create(
                classroom=classroom,
                date=date,
                defaults={
                    'marked_by': marked_by_user,
                    'status': 'under_review',
                    'submitted_at': timezone.now(),
                    'submitted_by': user,
                    'organization': org
                }
            )
            
            # If attendance already exists, update status to under_review
            if not created:
                # Update organization if missing
                if not attendance.organization and org:
                    attendance.organization = org
                attendance.status = 'under_review'
                attendance.submitted_at = timezone.now()
                attendance.submitted_by = user
                attendance.marked_by = marked_by_user
                attendance.save()
            else:
                # Save first time (already saved by get_or_create)
                pass
            
            # Clear existing student attendance records
            attendance.student_attendances.all().delete()
            
            # Create new student attendance records
            for student_data in student_attendance_data:
                StudentAttendance.objects.create(
                    attendance=attendance,
                    student_id=student_data['student_id'],
                    status=student_data['status'],
                    remarks=student_data.get('remarks', ''),
                    organization=org
                )
            
            # Update attendance summary
            attendance.update_counts()
            
            # Save attendance with updated status
            attendance.save()

            # Audit trail for marking. Previously only a notification fired, so
            # the Logs view had no record of who marked a class or when — only
            # approvals showed. organization is set so it is visible to the
            # org-scoped Logs viewers (AuditLog uses OrganizationManager).
            try:
                from .models import AuditLog
                AuditLog.objects.create(
                    feature='attendance',
                    action='mark',
                    entity_type='Attendance',
                    entity_id=attendance.id,
                    user=request.user,
                    organization=attendance.organization,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    changes={
                        'status': attendance.status,
                        'classroom': str(attendance.classroom),
                        'date': str(attendance.date),
                    },
                )
            except Exception as log_error:
                print(f"[WARN] Failed to write mark audit log: {log_error}")

            # Add edit history after saving
            attendance.add_edit_history(request.user, 'marked', 'Attendance marked and submitted for review')

            # Trigger consecutive absence alerts for class teacher
            try:
                alerts = process_consecutive_absence_alerts(attendance)
                if alerts:
                    print(f"[INFO] Consecutive absence alerts generated: {[alert.student_name for alert in alerts]}")
            except Exception as alert_error:
                print(f"[WARN] Failed to process consecutive absence alerts: {alert_error}")

            # Send notification to coordinator
            try:
                # Get coordinator for this classroom's level
                coordinator = None
                coordinator_user = None
                
                if classroom.grade and classroom.grade.level:
                    from coordinator.models import Coordinator
                    from django.contrib.auth import get_user_model
                    User = get_user_model()
                    
                    # Find coordinator for this level (considering shift)
                    coordinators = Coordinator.objects.filter(
                        is_currently_active=True
                    )
                    
                    # Check if coordinator manages this level
                    for coord in coordinators:
                        if coord.shift == 'both':
                            if coord.assigned_levels.exists():
                                if classroom.grade.level in coord.assigned_levels.all():
                                    coordinator = coord
                                    break
                            elif coord.level == classroom.grade.level:
                                coordinator = coord
                                break
                        else:
                            if coord.level == classroom.grade.level:
                                coordinator = coord
                                break
                    
                    # If no coordinator found, try to get from teacher's assigned coordinators
                    if not coordinator and teacher:
                        assigned_coords = teacher.assigned_coordinators.filter(is_currently_active=True).first()
                        if assigned_coords:
                            coordinator = assigned_coords
                    
                    # Get user for coordinator (by email or employee_code)
                    if coordinator:
                        try:
                            # Try by employee_code first
                            if coordinator.employee_code:
                                coordinator_user = User.objects.filter(username=coordinator.employee_code).first()
                            # Fallback to email
                            if not coordinator_user and coordinator.email:
                                coordinator_user = User.objects.filter(email=coordinator.email).first()
                        except Exception as user_error:
                            print(f"[WARN] Error finding user for coordinator {coordinator.full_name}: {user_error}")
                    
                    if coordinator and coordinator_user:
                        teacher_name = teacher.full_name if teacher else request.user.get_full_name() or request.user.username
                        classroom_name = str(classroom)
                        verb = f"Class teacher {teacher_name} has marked attendance"
                        target_text = f"for {classroom_name}. Please review the attendance."
                        
                        create_notification(
                            recipient=coordinator_user,
                            actor=request.user,
                            verb=verb,
                            target_text=target_text,
                            data={
                                'attendance_id': attendance.id,
                                'classroom_id': classroom.id,
                                'classroom_name': classroom_name,
                                'date': str(attendance.date),
                                'teacher_name': teacher_name
                            }
                        )
                        print(f"[OK] Sent attendance notification to coordinator {coordinator.full_name} (user: {coordinator_user.email})")
                    elif coordinator:
                        print(f"[WARN] Coordinator {coordinator.full_name} found but no user account exists (email: {coordinator.email}, employee_code: {coordinator.employee_code})")
                    else:
                        print(f"[WARN] No coordinator found for classroom {classroom_name} (level: {classroom.grade.level.name if classroom.grade and classroom.grade.level else 'N/A'})")
            except Exception as notif_error:
                print(f"[WARN] Failed to send attendance notification: {notif_error}")
                import traceback
                print(f"[WARN] Traceback: {traceback.format_exc()}")
                # Don't fail the attendance marking if notification fails
            
        return Response({
            'message': 'Attendance marked successfully',
            'attendance_id': attendance.id,
            'total_students': attendance.total_students,
            'present_count': attendance.present_count,
            'absent_count': attendance.absent_count,
            'late_count': attendance.late_count
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_bulk_attendance(request):
    """
    Mark attendance for entire class with simple present/absent status
    """
    try:
        classroom_id = request.data.get('classroom_id')
        date_str = request.data.get('date')
        student_attendance_data = request.data.get('student_attendance', [])
        
        if not classroom_id or not date_str:
            return Response({
                'error': 'classroom_id and date are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Convert date string to date object
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({
                'error': 'Invalid date format. Use YYYY-MM-DD.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        classroom = get_object_or_404(ClassRoom, id=classroom_id)
        
        # Check if date is a holiday (support multiple levels and grade-specific)
        from .models import Holiday
        level = classroom.grade.level if classroom.grade else None
        grade = classroom.grade if classroom.grade else None
        
        if level:
            # Check for holidays matching this level and date
            holidays = Holiday.objects.filter(
                date=date_obj
            ).filter(
                Q(levels=level) | Q(level=level)  # Support both old and new fields
            ).distinct()
            
            # Check if any holiday applies to this classroom
            for holiday in holidays:
                is_holiday = False
                # If holiday has specific grades, check if this classroom's grade is included
                if holiday.grades.exists():
                    if grade and grade in holiday.grades.all():
                        is_holiday = True
                else:
                    # No specific grades - applies to all grades in the level
                    is_holiday = True
                
                if is_holiday:
                    try:
                        is_teacher = request.user.is_teacher()
                    except Exception:
                        is_teacher = False
                    if is_teacher and not request.user.is_superuser:
                        return Response({
                            'error': f'This date is a holiday: {holiday.reason}. Attendance marking is disabled.',
                            'is_holiday': True,
                            'holiday_reason': holiday.reason
                        }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if it's a Sunday and auto-create weekend entry, and block teacher marking
        if date_obj.weekday() == 6:  # Sunday is 6 in Python's weekday()
            level = classroom.grade.level
            Weekend.objects.get_or_create(
                date=date_obj,
                level=level,
                defaults={'created_by': request.user}
            )
            # Teachers should not be able to mark Sunday attendance
            try:
                is_teacher = request.user.is_teacher()
            except Exception:
                is_teacher = False
            if is_teacher and not request.user.is_superuser:
                return Response({
                    'error': 'Weekend (Sunday): attendance marking is disabled',
                    'is_weekend': True
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get all students in this class (non-deleted students)
        # Removed is_active filter to ensure all students appear consistently
        all_students = Student.objects.filter(classroom=classroom, is_deleted=False)
        all_student_ids = list(all_students.values_list('id', flat=True))
        
        with transaction.atomic():
            # Get teacher from request user
            teacher = None
            try:
                teacher = Teacher.objects.get(employee_code=request.user.username)
            except Teacher.DoesNotExist:
                pass
            
            # Automatically set organization if user belongs to one
            org = None
            if not request.user.is_superadmin() and request.user.organization:
                org = request.user.organization
            
            # Create or get attendance record
            # marked_by must be a User instance, not Teacher
            if teacher and hasattr(teacher, 'user') and teacher.user is not None:
                marked_by_user = teacher.user
            else:
                marked_by_user = request.user
            attendance, created = Attendance.objects.get_or_create(
                classroom=classroom,
                date=date_obj,
                defaults={
                    'marked_by': marked_by_user,
                    'status': 'under_review',
                    'submitted_at': timezone.now(),
                    'submitted_by': request.user,
                    'organization': org
                }
            )
            
            # If attendance already exists, update status to under_review
            if not created:
                # Update organization if missing
                if not attendance.organization and org:
                    attendance.organization = org
                attendance.status = 'under_review'
                attendance.submitted_at = timezone.now()
                attendance.submitted_by = request.user
                attendance.marked_by = marked_by_user
            
            # Clear existing student attendance records
            attendance.student_attendances.all().delete()
            
            # Create student attendance records
            for student_data in student_attendance_data:
                student_id = student_data.get('student_id')
                attendance_status = student_data.get('status', 'present')
                remarks = student_data.get('remarks', '')
                
                if not student_id:
                    continue
                
                # Verify student belongs to this classroom
                try:
                    student = Student.objects.get(id=student_id, classroom=classroom)
                    StudentAttendance.objects.create(
                        attendance=attendance,
                        student=student,
                        status=attendance_status,
                        remarks=remarks,
                        created_by=request.user,
                        updated_by=request.user,
                        organization=attendance.organization
                    )
                except Student.DoesNotExist:
                    continue
            
            # Update attendance summary
            attendance.update_counts()
            
            # Save attendance with updated status
            attendance.save()

            # Audit trail for marking. Previously only a notification fired, so
            # the Logs view had no record of who marked a class or when — only
            # approvals showed. organization is set so it is visible to the
            # org-scoped Logs viewers (AuditLog uses OrganizationManager).
            try:
                from .models import AuditLog
                AuditLog.objects.create(
                    feature='attendance',
                    action='mark',
                    entity_type='Attendance',
                    entity_id=attendance.id,
                    user=request.user,
                    organization=attendance.organization,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    changes={
                        'status': attendance.status,
                        'classroom': str(attendance.classroom),
                        'date': str(attendance.date),
                    },
                )
            except Exception as log_error:
                print(f"[WARN] Failed to write mark audit log: {log_error}")

            # Add edit history after saving
            attendance.add_edit_history(request.user, 'marked', 'Attendance marked and submitted for review')
            
            # Trigger consecutive absence alerts for class teacher
            try:
                alerts = process_consecutive_absence_alerts(attendance)
                if alerts:
                    print(f"[INFO] Consecutive absence alerts generated: {[alert.student_name for alert in alerts]}")
            except Exception as alert_error:
                print(f"[WARN] Failed to process consecutive absence alerts: {alert_error}")

            # Send notification to coordinator
            try:
                # Teacher already retrieved above
                
                # Get coordinator for this classroom's level
                coordinator = None
                coordinator_user = None
                
                if classroom.grade and classroom.grade.level:
                    from coordinator.models import Coordinator
                    from django.contrib.auth import get_user_model
                    User = get_user_model()
                    
                    # Find coordinator for this level (considering shift)
                    coordinators = Coordinator.objects.filter(
                        is_currently_active=True
                    )
                    
                    # Check if coordinator manages this level
                    for coord in coordinators:
                        if coord.shift == 'both':
                            if coord.assigned_levels.exists():
                                if classroom.grade.level in coord.assigned_levels.all():
                                    coordinator = coord
                                    break
                            elif coord.level == classroom.grade.level:
                                coordinator = coord
                                break
                        else:
                            if coord.level == classroom.grade.level:
                                coordinator = coord
                                break
                    
                    # If no coordinator found, try to get from teacher's assigned coordinators
                    if not coordinator and teacher:
                        assigned_coords = teacher.assigned_coordinators.filter(is_currently_active=True).first()
                        if assigned_coords:
                            coordinator = assigned_coords
                    
                    # Get user for coordinator (by email or employee_code)
                    if coordinator:
                        try:
                            # Try by employee_code first
                            if coordinator.employee_code:
                                coordinator_user = User.objects.filter(username=coordinator.employee_code).first()
                            # Fallback to email
                            if not coordinator_user and coordinator.email:
                                coordinator_user = User.objects.filter(email=coordinator.email).first()
                        except Exception as user_error:
                            print(f"[WARN] Error finding user for coordinator {coordinator.full_name}: {user_error}")
                    
                    if coordinator and coordinator_user:
                        teacher_name = teacher.full_name if teacher else request.user.get_full_name() or request.user.username
                        classroom_name = str(classroom)
                        verb = f"Class teacher {teacher_name} has marked attendance"
                        target_text = f"for {classroom_name}. Please review the attendance."
                        
                        create_notification(
                            recipient=coordinator_user,
                            actor=request.user,
                            verb=verb,
                            target_text=target_text,
                            data={
                                'attendance_id': attendance.id,
                                'classroom_id': classroom.id,
                                'classroom_name': classroom_name,
                                'date': str(attendance.date),
                                'teacher_name': teacher_name
                            }
                        )
                        print(f"[OK] Sent attendance notification to coordinator {coordinator.full_name} (user: {coordinator_user.email})")
                    elif coordinator:
                        print(f"[WARN] Coordinator {coordinator.full_name} found but no user account exists (email: {coordinator.email}, employee_code: {coordinator.employee_code})")
                    else:
                        print(f"[WARN] No coordinator found for classroom {classroom_name} (level: {classroom.grade.level.name if classroom.grade and classroom.grade.level else 'N/A'})")
            except Exception as notif_error:
                print(f"[WARN] Failed to send attendance notification: {notif_error}")
                import traceback
                print(f"[WARN] Traceback: {traceback.format_exc()}")
                # Don't fail the attendance marking if notification fails
            
        return Response({
            'message': 'Bulk attendance marked successfully',
            'attendance_id': attendance.id,
            'total_students': attendance.total_students,
            'present_count': attendance.present_count,
            'absent_count': attendance.absent_count,
            'leave_count': attendance.leave_count,
            'attendance_percentage': attendance.attendance_percentage
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_class_attendance(request, classroom_id):
    """
    Get attendance records for a specific class
    """
    classroom = get_object_or_404(ClassRoom, id=classroom_id)
    date_param = request.GET.get('date')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if date_param:
        # Get attendance for specific date
        attendance = Attendance.objects.filter(
            classroom=classroom,
            date=date_param
        ).prefetch_related('student_attendances__student').first()
        if attendance:
            serializer = AttendanceSerializer(attendance, context={'request': request})
            return Response(serializer.data)
        else:
            return Response({'message': 'No attendance found for this date'})
    elif start_date or end_date:
        # Get attendance for date range
        attendance_records = Attendance.objects.filter(
            classroom=classroom
        ).prefetch_related('student_attendances__student')
        
        if start_date:
            attendance_records = attendance_records.filter(date__gte=start_date)
        if end_date:
            attendance_records = attendance_records.filter(date__lte=end_date)
            
        attendance_records = attendance_records.order_by('-date')
        serializer = AttendanceSerializer(attendance_records, many=True, context={'request': request})
        return Response(serializer.data)
    else:
        # Get all attendance records for the class
        attendance_records = Attendance.objects.filter(
            classroom=classroom
        ).prefetch_related('student_attendances__student').order_by('-date')
        serializer = AttendanceSerializer(attendance_records, many=True, context={'request': request})
        return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_attendance(request, student_id):
    """
    Get attendance history for a specific student
    """
    student = get_object_or_404(Student, id=student_id)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    attendance_records = StudentAttendance.objects.filter(
        student=student
    ).select_related('attendance')
    
    if start_date:
        attendance_records = attendance_records.filter(
            attendance__date__gte=start_date
        )
    if end_date:
        attendance_records = attendance_records.filter(
            attendance__date__lte=end_date
        )
    
    serializer = StudentAttendanceSerializer(attendance_records, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_class_students(request, classroom_id):
    """
    Get all students in a specific classroom
    """
    classroom = get_object_or_404(ClassRoom, id=classroom_id)
    
    # Check permissions - teacher can only see their assigned classes (supports multiple)
    user = request.user
    if user.is_teacher():
        try:
            # Find teacher by employee code (username)
            from teachers.models import Teacher
            teacher = Teacher.objects.get(employee_code=user.username)
            # allow if legacy single matches OR included in M2M assigned_classrooms OR classroom.class_teacher is this teacher
            allowed = False
            if teacher.assigned_classroom == classroom:
                allowed = True
            elif teacher.assigned_classrooms.filter(id=classroom.id).exists():
                allowed = True
            elif getattr(classroom, 'class_teacher_id', None) == teacher.id:
                allowed = True
            if not allowed:
                return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher profile not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Filter students: only non-deleted students should appear in attendance
    # Removed is_active filter to match student list view - all non-deleted students should be visible
    students = Student.objects.filter(classroom=classroom, is_deleted=False).order_by('name')
    
    student_data = []
    for student in students:
        student_data.append({
            'id': student.id,
            'name': student.name,
            'father_name': student.father_name,
            'father_cnic': student.father_cnic,
            'student_code': student.student_code,
            'photo': student.photo.url if student.photo else None,
            'gr_no': student.gr_no,
            'gender': student.gender,
            'student_id': student.student_id
        })
    
    return Response(student_data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_attendance_summary(request, classroom_id):
    """
    Get attendance summary for a classroom
    """
    classroom = get_object_or_404(ClassRoom, id=classroom_id)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        start_date = (timezone.now().date() - timedelta(days=30))
    if not end_date:
        end_date = timezone.now().date()
    
    attendance_records = Attendance.objects.filter(
        classroom=classroom,
        date__range=[start_date, end_date]
    ).order_by('-date')
    
    summary_data = []
    for attendance in attendance_records:
        attendance_percentage = attendance.attendance_percentage
        
        summary_data.append({
            'classroom_id': classroom.id,
            'classroom_name': str(classroom),
            'date': attendance.date,
            'total_students': attendance.total_students,
            'present_count': attendance.present_count,
            'absent_count': attendance.absent_count,
            'late_count': attendance.late_count,
            'leave_count': attendance.leave_count,  # Add leave_count to response
            'attendance_percentage': round(attendance_percentage, 2)
        })
    
    return Response(summary_data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_teacher_classes(request):
    """
    Get all classes assigned to the current teacher
    """
    try:
        user = request.user
        
        # Find teacher by employee code (username) since there's no direct relationship
        from teachers.models import Teacher
        try:
            teacher = Teacher.objects.get(employee_code=user.username)
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if not teacher:
            return Response({'error': 'Teacher profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get classes where teacher is class teacher
        classrooms = ClassRoom.objects.filter(class_teacher=teacher)
        
        class_data = []
        for classroom in classrooms:
            class_data.append({
                'id': classroom.id,
                'name': str(classroom),
                'code': classroom.code,
                'grade': classroom.grade.name,
                'section': classroom.section,
                'shift': classroom.shift,
                'campus': classroom.grade.level.campus.campus_name if classroom.grade.level.campus else None
            })
        
        return Response(class_data)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_attendance(request, attendance_id):
    """
    Edit existing attendance record
    Teachers can edit within 7 days, Coordinators can edit anytime for their level
    """
    try:
        attendance = get_object_or_404(Attendance, id=attendance_id, is_deleted=False)
        user = request.user

        # Check if the attendance date is a holiday
        classroom = attendance.classroom
        level = classroom.grade.level if classroom.grade else None
        grade = classroom.grade if classroom.grade else None
        if level:
            from .models import Holiday
            holidays = Holiday.objects.filter(
                date=attendance.date
            ).filter(
                Q(levels=level) | Q(level=level)
            ).distinct()
            for holiday in holidays:
                if holiday.grades.exists():
                    if grade and grade in holiday.grades.all():
                        return Response({
                            'error': f'Cannot edit attendance on a holiday: {holiday.reason}.',
                            'is_holiday': True,
                            'holiday_reason': holiday.reason
                        }, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({
                        'error': f'Cannot edit attendance on a holiday: {holiday.reason}.',
                        'is_holiday': True,
                        'holiday_reason': holiday.reason
                    }, status=status.HTTP_400_BAD_REQUEST)

        # Check if user can edit this attendance
        can_edit = False
        edit_reason = None
        
        # SuperAdmin can edit anything
        if user.is_superuser:
            can_edit = True
            edit_reason = "SuperAdmin edit"
        
        # Check teacher permissions (7-day limit)
        elif user.is_teacher():
            try:
                # Find teacher by employee code (username) since there's no direct relationship
                from teachers.models import Teacher
                teacher = Teacher.objects.get(employee_code=user.username)
                is_allowed = False
                if teacher and teacher.assigned_classroom == attendance.classroom:
                    is_allowed = True
                elif teacher and teacher.assigned_classrooms.filter(id=attendance.classroom_id).exists():
                    is_allowed = True
                elif teacher and getattr(attendance.classroom, 'class_teacher_id', None) == teacher.id:
                    is_allowed = True

                if is_allowed:
                    # Teachers can edit under_review attendance within 7 days
                    if attendance.status == 'under_review':
                        days_diff = (timezone.now().date() - attendance.date).days
                        if days_diff <= 7:
                            can_edit = True
                            edit_reason = "Teacher edit within 7 days"
                        else:
                            return Response({
                                'error': f'Cannot edit attendance older than 7 days. This attendance is {days_diff} days old.'
                            }, status=status.HTTP_403_FORBIDDEN)
                    elif attendance.status == 'approved':
                        return Response({
                            'error': 'Cannot edit approved attendance. Please contact your coordinator if changes are needed.'
                        }, status=status.HTTP_403_FORBIDDEN)
                    else:
                        # For draft or submitted (legacy), allow edit but convert to under_review
                        days_diff = (timezone.now().date() - attendance.date).days
                        if days_diff <= 7:
                            can_edit = True
                            edit_reason = "Teacher edit within 7 days"
                        else:
                            return Response({
                                'error': f'Cannot edit attendance older than 7 days. This attendance is {days_diff} days old.'
                            }, status=status.HTTP_403_FORBIDDEN)
            except Teacher.DoesNotExist:
                pass
        
        # Check coordinator permissions (unlimited time for their level)
        elif user.is_coordinator():
            # Coordinator can edit attendance for their managed levels (no 7-day limit)
            from coordinator.models import Coordinator
            coordinator = Coordinator.get_for_user(user)
            if not coordinator or not coordinator.is_currently_active:
                return Response({'error': 'Coordinator profile not found'}, status=status.HTTP_404_NOT_FOUND)

            # Support coordinators with multiple assigned levels and 'both' shifts
            allowed = False
            if coordinator.shift == 'both':
                if hasattr(coordinator, 'assigned_levels') and coordinator.assigned_levels.exists():
                    if attendance.classroom.grade.level in coordinator.assigned_levels.all():
                        allowed = True
                elif coordinator.level:
                    if attendance.classroom.grade.level == coordinator.level:
                        allowed = True
            else:
                if coordinator.level and attendance.classroom.grade.level == coordinator.level:
                    allowed = True

            if allowed:
                can_edit = True
                edit_reason = "Coordinator edit"
        
        # Check principal permissions
        elif user.is_principal():
            try:
                # Find principal by email since there's no direct relationship
                from principals.models import Principal
                principal = Principal.objects.get(email=user.email)
                if (principal and principal.is_currently_active and 
                    principal.campus == attendance.classroom.campus):
                    can_edit = True
                    edit_reason = "Principal edit"
            except Principal.DoesNotExist:
                pass
        
        if not can_edit:
            return Response({
                'error': 'You do not have permission to edit this attendance'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get new attendance data
        data = request.data
        student_attendance_data = data.get('student_attendance', [])
        
        if not student_attendance_data:
            return Response({
                'error': 'Student attendance data is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        with transaction.atomic():
            # Clear existing student attendance records
            attendance.student_attendances.all().delete()
            
            # Create new student attendance records
            for student_data in student_attendance_data:
                student_id = student_data.get('student_id')
                attendance_status = student_data.get('status', 'present')
                remarks = student_data.get('remarks', '')
                
                if not student_id:
                    continue
                
                try:
                    student = Student.objects.get(id=student_id, classroom=attendance.classroom)
                    StudentAttendance.objects.create(
                        student=student,
                        attendance=attendance,
                        status=attendance_status,
                        remarks=remarks,
                        created_by=user,
                        updated_by=user,
                        organization=attendance.organization
                    )
                except Student.DoesNotExist:
                    continue
            
            # Update attendance counts
            attendance.update_counts()
            
            # Add edit history
            attendance.add_edit_history(
                user=user,
                action='edited',
                reason=edit_reason,
                changes={
                    'edited_at': timezone.now().isoformat(),
                    'student_count': len(student_attendance_data)
                }
            )
            
            # Update status and marked_by if it's a teacher
            teacher = None
            if user.is_teacher():
                # Keep status as under_review if not approved
                if attendance.status != 'approved':
                    attendance.status = 'under_review'
                    attendance.submitted_at = timezone.now()
                    attendance.submitted_by = user
                
                # Get teacher for notification
                # marked_by must be a User instance, not Teacher
                try:
                    teacher = Teacher.objects.get(employee_code=user.username)
                    # Use teacher.user if available, otherwise use user
                    if hasattr(teacher, 'user') and teacher.user is not None:
                        attendance.marked_by = teacher.user
                    else:
                        attendance.marked_by = user
                except Teacher.DoesNotExist:
                    print(f"[WARN] Teacher not found for user {user.username}")
                    attendance.marked_by = user
                    pass
                
                # Save only specific fields to avoid updating created_at
                attendance.save(update_fields=['marked_by', 'status', 'submitted_at', 'submitted_by', 'updated_at'])
            
        # Trigger consecutive absence alerts for class teacher
        try:
            alerts = process_consecutive_absence_alerts(attendance)
            if alerts:
                print(f"[INFO] Consecutive absence alerts generated: {[alert.student_name for alert in alerts]}")
        except Exception as alert_error:
            print(f"[WARN] Failed to process consecutive absence alerts: {alert_error}")

            # Send notification to coordinator when teacher updates attendance
            # (Don't send if coordinator is editing their own attendance)
            if user.is_teacher():
                print(f"[DEBUG] Teacher updating attendance - user: {user.username}, is_teacher: {user.is_teacher()}, teacher: {teacher}")
                try:
                    # Get coordinator for this classroom's level
                    coordinator = None
                    coordinator_user = None
                    classroom = attendance.classroom
                    
                    print(f"[DEBUG] Classroom: {classroom}, Grade: {classroom.grade if classroom.grade else 'None'}, Level: {classroom.grade.level if classroom.grade and classroom.grade.level else 'None'}")
                    
                    if classroom.grade and classroom.grade.level:
                        from coordinator.models import Coordinator
                        from django.contrib.auth import get_user_model
                        User = get_user_model()
                        
                        # Find coordinator for this level (considering shift)
                        coordinators = Coordinator.objects.filter(
                            is_currently_active=True
                        )
                        
                        print(f"[DEBUG] Found {coordinators.count()} active coordinators")
                        
                        # Check if coordinator manages this level
                        for coord in coordinators:
                            print(f"[DEBUG] Checking coordinator {coord.full_name} - shift: {coord.shift}, level: {coord.level}, assigned_levels: {list(coord.assigned_levels.all()) if coord.assigned_levels.exists() else 'None'}")
                            if coord.shift == 'both':
                                if coord.assigned_levels.exists():
                                    if classroom.grade.level in coord.assigned_levels.all():
                                        coordinator = coord
                                        break
                                elif coord.level == classroom.grade.level:
                                    coordinator = coord
                                    break
                            else:
                                if coord.level == classroom.grade.level:
                                    coordinator = coord
                                    break
                        
                        # If no coordinator found, try to get from teacher's assigned coordinators
                        if not coordinator and teacher:
                            print(f"[DEBUG] No coordinator found by level, trying teacher's assigned coordinators")
                            assigned_coords = teacher.assigned_coordinators.filter(is_currently_active=True).first()
                            if assigned_coords:
                                coordinator = assigned_coords
                                print(f"[DEBUG] Found coordinator from teacher's assigned coordinators: {coordinator.full_name}")
                        
                        # Get user for coordinator (by email or employee_code)
                        if coordinator:
                            try:
                                # Try by employee_code first
                                if coordinator.employee_code:
                                    coordinator_user = User.objects.filter(username=coordinator.employee_code).first()
                                # Fallback to email
                                if not coordinator_user and coordinator.email:
                                    coordinator_user = User.objects.filter(email=coordinator.email).first()
                            except Exception as user_error:
                                print(f"[WARN] Error finding user for coordinator {coordinator.full_name}: {user_error}")
                        
                        if coordinator and coordinator_user:
                            teacher_name = teacher.full_name if teacher else user.get_full_name() or user.username
                            classroom_name = str(classroom)
                            verb = f"Class teacher {teacher_name} has updated attendance"
                            target_text = f"for {classroom_name}. Please review the attendance."
                            
                            create_notification(
                                recipient=coordinator_user,
                                actor=user,
                                verb=verb,
                                target_text=target_text,
                                data={
                                    'attendance_id': attendance.id,
                                    'classroom_id': classroom.id,
                                    'classroom_name': classroom_name,
                                    'date': str(attendance.date),
                                    'teacher_name': teacher_name,
                                    'action': 'updated'
                                }
                            )
                            print(f"[OK] Sent attendance update notification to coordinator {coordinator.full_name} (user: {coordinator_user.email})")
                        elif coordinator:
                            print(f"[WARN] Coordinator {coordinator.full_name} found but no user account exists (email: {coordinator.email}, employee_code: {coordinator.employee_code})")
                        else:
                            print(f"[WARN] No coordinator found for classroom {classroom_name} (level: {classroom.grade.level.name if classroom.grade and classroom.grade.level else 'N/A'})")
                except Exception as notif_error:
                    print(f"[WARN] Failed to send attendance update notification: {notif_error}")
                    import traceback
                    print(f"[WARN] Traceback: {traceback.format_exc()}")
                    # Don't fail the attendance update if notification fails
        
        # Return updated attendance data
        serializer = AttendanceSerializer(attendance, context={'request': request})
        return Response({
            'message': 'Attendance updated successfully',
            'attendance': serializer.data
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_attendance_for_date(request, classroom_id, date):
    """
    Get attendance for a specific date
    """
    try:
        classroom = get_object_or_404(ClassRoom, id=classroom_id)
        user = request.user
        
        # Check permissions (support multi-class teachers)
        if user.is_teacher():
            try:
                # Find teacher by employee code (username) since there's no direct relationship
                from teachers.models import Teacher
                teacher = Teacher.objects.get(employee_code=user.username)
                allowed = False
                if teacher.assigned_classroom == classroom:
                    allowed = True
                elif teacher.assigned_classrooms.filter(id=classroom.id).exists():
                    allowed = True
                elif getattr(classroom, 'class_teacher_id', None) == teacher.id:
                    allowed = True
                if not allowed:
                    return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
            except Teacher.DoesNotExist:
                return Response({'error': 'Teacher profile not found'}, status=status.HTTP_404_NOT_FOUND)
        elif user.is_coordinator():
            # Coordinator can access attendance for classrooms in their managed levels
            from coordinator.models import Coordinator
            coordinator = Coordinator.get_for_user(user)
            if not coordinator:
                return Response({'error': 'Coordinator profile not found'}, status=status.HTTP_404_NOT_FOUND)

            # Check if classroom is in coordinator's managed levels
            allowed = False
            # Always check assigned_levels first (supports multi-level coordinators regardless of shift).
            # Fallback to single 'level' field for older coordinator records.
            if coordinator.assigned_levels.exists():
                if classroom.grade.level in coordinator.assigned_levels.all():
                    allowed = True
            elif coordinator.level:
                if classroom.grade.level == coordinator.level:
                    allowed = True

            if not allowed:
                return Response({'error': 'Access denied - Classroom not in your managed levels'}, status=status.HTTP_403_FORBIDDEN)
        
        # Auto-create weekend entry for Sundays (no attendance records should be created)
        try:
            from datetime import datetime as _dt
            date_obj = _dt.strptime(date, '%Y-%m-%d').date()
            if date_obj.weekday() == 6:  # Sunday
                Weekend.objects.get_or_create(
                    date=date_obj,
                    level=classroom.grade.level,
                    defaults={'created_by': request.user}
                )
        except Exception:
            pass

        # Get attendance for the date
        try:
            attendance = Attendance.objects.get(
                classroom=classroom,
                date=date,
                is_deleted=False
            )
            
            # Get student attendance records
            student_attendances = attendance.student_attendances.all()
            
            # Use serializer to get display_status
            serializer = AttendanceSerializer(attendance, context={'request': request})
            serializer_data = serializer.data
            
            attendance_data = {
                'id': attendance.id,
                'date': attendance.date.isoformat(),
                'classroom': {
                    'id': classroom.id,
                    'name': str(classroom),
                    'code': classroom.code
                },
                'total_students': attendance.total_students,
                'present_count': attendance.present_count,
                'absent_count': attendance.absent_count,
                'late_count': attendance.late_count,
                'leave_count': attendance.leave_count,
                'attendance_percentage': attendance.attendance_percentage,
                'is_editable': attendance.is_editable,
                'marked_at': attendance.marked_at.isoformat(),
                'marked_by': attendance.marked_by.get_full_name() if attendance.marked_by else None,
                'status': attendance.status,
                'display_status': serializer_data.get('display_status', attendance.status),
                'student_attendance': [
                    {
                        'student_id': sa.student.id,
                        'student_name': sa.student.name,
                        'student_code': sa.student.student_code or sa.student.student_id or sa.student.gr_no or f"ID-{sa.student.id}",
                        'student_gender': sa.student.gender,
                        'status': sa.status,
                        'remarks': sa.remarks or ''
                    }
                    for sa in student_attendances
                ],
                'edit_history': attendance.update_history
            }
            
            return Response(attendance_data)
            
        except Attendance.DoesNotExist:
            # Also tell client if the date is a weekend
            from datetime import datetime as _dt
            is_weekend = False
            try:
                _d = _dt.strptime(date, '%Y-%m-%d').date()
                is_weekend = (_d.weekday() == 6)
            except Exception:
                pass
            return Response({
                'message': 'No attendance found for this date',
                'date': date,
                'classroom_id': classroom_id,
                'is_weekend': is_weekend
            })
            
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_coordinator_classes(request):
    """
    Get all classes in coordinator's assigned level
    """
    try:
        user = request.user
        
        if not user.is_coordinator():
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Find coordinator by user (robust lookup)
        from coordinator.models import Coordinator
        coordinator = Coordinator.get_for_user(user)
        if not coordinator or not coordinator.is_currently_active:
            return Response({'error': 'Coordinator profile not found or inactive'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get all classes in coordinator's level(s)
        managed_levels = []
        if coordinator.assigned_levels.exists():
            managed_levels = list(coordinator.assigned_levels.all())
        if not managed_levels and coordinator.level:
            managed_levels = [coordinator.level]
        if not managed_levels:
            return Response({'error': 'No level assigned to coordinator'}, status=status.HTTP_404_NOT_FOUND)
        
        classrooms = ClassRoom.objects.filter(
            grade__level__in=managed_levels
        ).select_related('grade', 'class_teacher', 'grade__level__campus')
        
        
        class_data = []
        for classroom in classrooms:
            # Include level information so frontend can build a level selection dropdown
            level_info = None
            try:
                lvl = classroom.grade.level
                level_info = {'id': lvl.id, 'name': lvl.name}
            except Exception:
                level_info = None

            class_data.append({
                'id': classroom.id,
                'name': str(classroom),  # This uses the __str__ method
                'code': classroom.code,
                'grade': classroom.grade.name,
                'section': classroom.section,
                'shift': classroom.shift,
                'level': level_info,
                'campus': classroom.grade.level.campus.campus_name if classroom.grade.level.campus else None,
                'class_teacher': {
                    'id': classroom.class_teacher.id if classroom.class_teacher else None,
                    'name': classroom.class_teacher.full_name if classroom.class_teacher else None,
                    'employee_code': classroom.class_teacher.employee_code if classroom.class_teacher else None
                } if classroom.class_teacher else None,
                'student_count': classroom.students.count()
            })
        
        return Response(class_data)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_level_attendance_summary(request, level_id):
    """
    Get attendance summary for all classes in a level
    """
    try:
        user = request.user
        
        # Check permissions
        if user.is_coordinator():
            coordinator = Coordinator.get_for_user(user)
            if not coordinator:
                return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)

            managed_level_ids = set()
            if hasattr(coordinator, "assigned_levels") and coordinator.assigned_levels.exists():
                managed_level_ids.update(coordinator.assigned_levels.values_list('id', flat=True))
            if coordinator.level_id:
                managed_level_ids.add(coordinator.level_id)

            if not managed_level_ids:
                return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)

            if level_id not in managed_level_ids:
                return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        elif user.is_principal():
            try:
                # Find principal by email since there's no direct relationship
                from principals.models import Principal
                principal = Principal.objects.get(email=user.email)
                if not principal or not principal.is_currently_active:
                    return Response({'error': 'Principal profile not found'}, status=status.HTTP_404_NOT_FOUND)
            except Principal.DoesNotExist:
                return Response({'error': 'Principal profile not found'}, status=status.HTTP_404_NOT_FOUND)
        elif not user.is_superuser:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get date range
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if not start_date:
            start_date = (timezone.now() - timedelta(days=30)).date()
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            
        if not end_date:
            end_date = timezone.now().date()
        else:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Get all classes in the level
        classrooms = ClassRoom.objects.filter(
            grade__level_id=level_id
        ).select_related('grade', 'grade__level__campus')
        
        summary_data = []
        total_students = 0
        total_present = 0
        total_absent = 0
        total_late = 0
        total_leave = 0
        
        for classroom in classrooms:
            # Get attendance records for this classroom in date range
            attendances = Attendance.objects.filter(
                classroom=classroom,
                date__range=[start_date, end_date]
            )
            
            
            classroom_total_students = classroom.students.count()
            classroom_present = sum(att.present_count for att in attendances)
            classroom_absent = sum(att.absent_count for att in attendances)
            classroom_late = sum(att.late_count for att in attendances)
            classroom_leave = sum(att.leave_count for att in attendances)
            classroom_records = attendances.count()
            
            avg_percentage = 0
            if classroom_records > 0:
                try:
                    # Safely get attendance percentages, handling None values
                    percentages = []
                    for att in attendances:
                        try:
                            pct = att.attendance_percentage
                            if pct is not None and not (isinstance(pct, float) and (pct != pct)):  # Check for NaN
                                percentages.append(pct)
                        except (ZeroDivisionError, TypeError, AttributeError):
                            continue
                    
                    if percentages:
                        total_percentage = sum(percentages)
                        avg_percentage = total_percentage / len(percentages)
                except (ZeroDivisionError, TypeError):
                    avg_percentage = 0
            
            summary_data.append({
                'classroom': {
                    'id': classroom.id,
                    'name': str(classroom),
                    'code': classroom.code,
                    'grade': classroom.grade.name,
                    'section': classroom.section,
                    'shift': classroom.shift,
                    'campus': classroom.grade.level.campus.campus_name if classroom.grade.level.campus else None
                },
                'student_count': classroom_total_students,
                'records_count': classroom_records,
                'total_present': classroom_present,
                'total_absent': classroom_absent,
                'total_late': classroom_late,
                'total_leave': classroom_leave,
                'average_percentage': round(avg_percentage, 2),
                'last_attendance': attendances.order_by('-date').first().date.isoformat() if attendances.exists() else None
            })
            
            total_students += classroom_total_students
            total_present += classroom_present
            total_absent += classroom_absent
            total_late += classroom_late
            total_leave += classroom_leave
        
        # Calculate overall statistics
        overall_percentage = 0
        try:
            total_attendance = total_present + total_absent
            if total_attendance > 0:
                overall_percentage = round((total_present / total_attendance) * 100, 2)
        except (ZeroDivisionError, TypeError):
            overall_percentage = 0
        
        return Response({
            'level_id': level_id,
            'date_range': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat()
            },
            'summary': {
                'total_classrooms': len(summary_data),
                'total_students': total_students,
                'total_present': total_present,
                'total_absent': total_absent,
                'total_late': total_late,
                'total_leave': total_leave,
                'overall_percentage': overall_percentage
            },
            'classrooms': summary_data
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def submit_attendance(request, attendance_id):
    """Teacher submits draft attendance for review"""
    try:
        attendance = get_object_or_404(Attendance, id=attendance_id, is_deleted=False)
        
        # Verify teacher can submit
        if attendance.status != 'draft':
            return Response({'error': 'Can only submit draft attendance'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Verify user is teacher of this class
        teacher = Teacher.objects.get(employee_code=request.user.username)
        if teacher.assigned_classroom != attendance.classroom:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        with transaction.atomic():
            attendance.status = 'submitted'
            attendance.submitted_at = timezone.now()
            attendance.submitted_by = request.user
            attendance.add_edit_history(request.user, 'submitted', 'Submitted for coordinator review')
            attendance.save()
            
            # Create audit log
            from .models import AuditLog
            AuditLog.objects.create(
                feature='attendance',
                action='submit',
                entity_type='Attendance',
                entity_id=attendance.id,
                user=request.user,
                ip_address=request.META.get('REMOTE_ADDR'),
                changes={'status': 'submitted'},
                reason='Submitted for coordinator review'
            )
        
        return Response({'message': 'Attendance submitted successfully'})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def review_attendance(request, attendance_id):
    """Coordinator moves attendance to under_review"""
    try:
        attendance = get_object_or_404(Attendance, id=attendance_id, is_deleted=False)
        
        if attendance.status != 'submitted':
            return Response({'error': 'Can only review submitted attendance'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Verify coordinator has access
        from coordinator.models import Coordinator
        coordinator = Coordinator.get_for_user(request.user)
        if not coordinator or not coordinator.is_currently_active:
            return Response({'error': 'Coordinator profile not found'}, status=status.HTTP_404_NOT_FOUND)

        # Support coordinators with multiple assigned levels and 'both' shifts
        allowed = False
        if coordinator.shift == 'both':
            if hasattr(coordinator, 'assigned_levels') and coordinator.assigned_levels.exists():
                if attendance.classroom.grade.level in coordinator.assigned_levels.all():
                    allowed = True
            elif coordinator.level:
                if attendance.classroom.grade.level == coordinator.level:
                    allowed = True
        else:
            if coordinator.level and attendance.classroom.grade.level == coordinator.level:
                allowed = True

        if not allowed:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        with transaction.atomic():
            attendance.status = 'under_review'
            attendance.reviewed_at = timezone.now()
            attendance.reviewed_by = request.user
            attendance.add_edit_history(request.user, 'review', 'Under coordinator review')
            attendance.save()
            
            from .models import AuditLog
            AuditLog.objects.create(
                feature='attendance',
                action='review',
                entity_type='Attendance',
                entity_id=attendance.id,
                user=request.user,
                ip_address=request.META.get('REMOTE_ADDR'),
                changes={'status': 'under_review'}
            )
        
        return Response({'message': 'Attendance moved to under review'})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def finalize_attendance(request, attendance_id):
    """Coordinator finalizes attendance (locks it)"""
    try:
        attendance = get_object_or_404(Attendance, id=attendance_id, is_deleted=False)
        
        if attendance.status not in ['draft', 'submitted', 'under_review']:
            return Response({'error': 'Can only finalize draft, submitted, or under_review attendance'}, status=status.HTTP_400_BAD_REQUEST)
        
        from coordinator.models import Coordinator
        coordinator = Coordinator.get_for_user(request.user)
        if not coordinator or not coordinator.is_currently_active:
            return Response({'error': 'Coordinator profile not found'}, status=status.HTTP_404_NOT_FOUND)

        # Support coordinators with multiple assigned levels and 'both' shifts
        allowed = False
        if coordinator.shift == 'both':
            if hasattr(coordinator, 'assigned_levels') and coordinator.assigned_levels.exists():
                if attendance.classroom.grade.level in coordinator.assigned_levels.all():
                    allowed = True
            elif coordinator.level:
                if attendance.classroom.grade.level == coordinator.level:
                    allowed = True
        else:
            if coordinator.level and attendance.classroom.grade.level == coordinator.level:
                allowed = True

        if not allowed:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        with transaction.atomic():
            attendance.status = 'approved'
            attendance.is_final = True
            attendance.finalized_at = timezone.now()
            attendance.finalized_by = request.user
            attendance.add_edit_history(request.user, 'finalize', 'Finalized by coordinator')
            attendance.save()

            from .models import AuditLog
            AuditLog.objects.create(
                feature='attendance',
                action='finalize',
                entity_type='Attendance',
                entity_id=attendance.id,
                user=request.user,
                ip_address=request.META.get('REMOTE_ADDR'),
                changes={'status': 'approved'}
            )
        
        return Response({'message': 'Attendance finalized successfully'})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def coordinator_approve_attendance(request, attendance_id):
    """Coordinator directly approves attendance (bypasses review step)"""
    try:
        attendance = get_object_or_404(Attendance, id=attendance_id, is_deleted=False)
        
        # Check if the attendance date is a holiday
        classroom = attendance.classroom
        level = classroom.grade.level if classroom.grade else None
        grade = classroom.grade if classroom.grade else None
        if level:
            from .models import Holiday
            holidays = Holiday.objects.filter(
                date=attendance.date
            ).filter(
                Q(levels=level) | Q(level=level)
            ).distinct()
            for holiday in holidays:
                if holiday.grades.exists():
                    if grade and grade in holiday.grades.all():
                        return Response({
                            'error': f'Cannot approve attendance on a holiday: {holiday.reason}.',
                            'is_holiday': True,
                            'holiday_reason': holiday.reason
                        }, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({
                        'error': f'Cannot approve attendance on a holiday: {holiday.reason}.',
                        'is_holiday': True,
                        'holiday_reason': holiday.reason
                    }, status=status.HTTP_400_BAD_REQUEST)

        # Check if attendance can be approved (draft, submitted, or under_review)
        if attendance.status not in ['draft', 'submitted', 'under_review']:
            return Response({'error': 'Can only approve draft, submitted or under review attendance'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Verify coordinator has access
        from coordinator.models import Coordinator
        coordinator = Coordinator.get_for_user(request.user)
        if not coordinator:
            return Response({'error': 'Coordinator profile not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check membership - always check assigned_levels first, fallback to level field
        allowed = False
        if coordinator.assigned_levels.exists():
            if attendance.classroom.grade.level in coordinator.assigned_levels.all():
                allowed = True
        elif coordinator.level:
            if coordinator.level == attendance.classroom.grade.level:
                allowed = True

        if not allowed:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        with transaction.atomic():
            # Move directly to approved status
            attendance.status = 'approved'
            attendance.is_final = True
            attendance.finalized_at = timezone.now()
            attendance.finalized_by = request.user
            attendance.add_edit_history(request.user, 'coordinator_approve', 'Directly approved by coordinator')
            attendance.save()

            from .models import AuditLog
            AuditLog.objects.create(
                feature='attendance',
                action='approve',
                entity_type='Attendance',
                entity_id=attendance.id,
                user=request.user,
                # organization is required for the log to be visible: AuditLog
                # uses OrganizationManager, so a null org is invisible to every
                # org-scoped viewer (org-admin, principal).
                organization=attendance.organization,
                ip_address=request.META.get('REMOTE_ADDR'),
                changes={'status': 'approved'}
            )
            
            # Send notification to teacher
            try:
                teacher_user = None
                teacher = None
                
                # Get teacher from marked_by or from classroom
                if attendance.marked_by:
                    teacher_user = attendance.marked_by
                    # Try to find teacher profile
                    try:
                        teacher = Teacher.objects.get(user=teacher_user)
                    except Teacher.DoesNotExist:
                        # Try by employee_code
                        try:
                            teacher = Teacher.objects.get(employee_code=teacher_user.username)
                        except Teacher.DoesNotExist:
                            pass
                elif attendance.classroom and attendance.classroom.class_teacher:
                    teacher = attendance.classroom.class_teacher
                    if teacher and teacher.user:
                        teacher_user = teacher.user
                
                if teacher_user:
                    coordinator_name = coordinator.full_name if coordinator else request.user.get_full_name() or request.user.username
                    classroom_name = str(attendance.classroom)
                    verb = "Your attendance has been approved"
                    target_text = f"by {coordinator_name} for {classroom_name} on {attendance.date.strftime('%B %d, %Y')}."
                    
                    create_notification(
                        recipient=teacher_user,
                        actor=request.user,
                        verb=verb,
                        target_text=target_text,
                        data={
                            'attendance_id': attendance.id,
                            'classroom_id': attendance.classroom.id,
                            'classroom_name': classroom_name,
                            'date': str(attendance.date),
                            'coordinator_name': coordinator_name,
                            'action': 'approved'
                        }
                    )
                    print(f"[OK] Sent approval notification to teacher {teacher.full_name if teacher else teacher_user.get_full_name()} (user: {teacher_user.email})")
                else:
                    print(f"[WARN] No teacher user found for attendance {attendance.id} (marked_by: {attendance.marked_by}, classroom: {attendance.classroom})")
            except Exception as notif_error:
                print(f"[WARN] Failed to send approval notification: {notif_error}")
                import traceback
                print(f"[WARN] Traceback: {traceback.format_exc()}")
                # Don't fail the approval if notification fails
        
        return Response({'message': 'Attendance approved successfully by coordinator'})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def coordinator_bulk_approve_attendance(request):
    """Coordinator bulk approves multiple attendances at once"""
    try:
        attendance_ids = request.data.get('attendance_ids', [])
        comment = request.data.get('comment', '')
        
        if not attendance_ids or not isinstance(attendance_ids, list):
            return Response({'error': 'attendance_ids must be a non-empty list'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Verify coordinator has access
        from coordinator.models import Coordinator
        coordinator = Coordinator.get_for_user(request.user)
        if not coordinator:
            return Response({'error': 'Coordinator profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        approved_count = 0
        failed_count = 0
        errors = []
        
        with transaction.atomic():
            for attendance_id in attendance_ids:
                try:
                    attendance = get_object_or_404(Attendance, id=attendance_id, is_deleted=False)

                    # Check if the attendance date is a holiday
                    classroom_obj = attendance.classroom
                    level_obj = classroom_obj.grade.level if classroom_obj.grade else None
                    grade_obj = classroom_obj.grade if classroom_obj.grade else None
                    if level_obj:
                        h_dates = Holiday.objects.filter(
                            date=attendance.date
                        ).filter(
                            Q(levels=level_obj) | Q(level=level_obj)
                        ).distinct()
                        is_holiday_blocked = False
                        for h_rec in h_dates:
                            if h_rec.grades.exists():
                                if grade_obj and grade_obj in h_rec.grades.all():
                                    is_holiday_blocked = True
                                    break
                            else:
                                is_holiday_blocked = True
                                break
                        if is_holiday_blocked:
                            failed_count += 1
                            errors.append(f"Attendance {attendance_id}: Cannot approve on a holiday")
                            continue

                    # Check if attendance can be approved
                    if attendance.status not in ['draft', 'submitted', 'under_review']:
                        failed_count += 1
                        errors.append(f"Attendance {attendance_id}: Cannot approve (status: {attendance.status})")
                        continue
                    
                    # Check membership (support assigned_levels for 'both' shifts)
                    allowed = False
                    if coordinator.assigned_levels.exists():
                        if attendance.classroom.grade.level in coordinator.assigned_levels.all():
                            allowed = True
                    elif coordinator.level:
                        if coordinator.level == attendance.classroom.grade.level:
                            allowed = True

                    if not allowed:
                        failed_count += 1
                        errors.append(f"Attendance {attendance_id}: Access denied")
                        continue
                    
                    # Approve attendance
                    attendance.status = 'approved'
                    attendance.is_final = True
                    attendance.finalized_at = timezone.now()
                    attendance.finalized_by = request.user
                    attendance.add_edit_history(request.user, 'coordinator_approve', f'Bulk approved by coordinator{": " + comment if comment else ""}')
                    attendance.save()
                    
                    # Create audit log
                    from .models import AuditLog
                    AuditLog.objects.create(
                        feature='attendance',
                        action='approve',
                        entity_type='Attendance',
                        entity_id=attendance.id,
                        user=request.user,
                        organization=attendance.organization,
                        ip_address=request.META.get('REMOTE_ADDR'),
                        changes={'status': 'approved', 'bulk_approval': True}
                    )
                    
                    # Send notification to teacher
                    try:
                        teacher_user = None
                        teacher = None
                        
                        if attendance.marked_by:
                            teacher_user = attendance.marked_by
                            try:
                                teacher = Teacher.objects.get(user=teacher_user)
                            except Teacher.DoesNotExist:
                                try:
                                    teacher = Teacher.objects.get(employee_code=teacher_user.username)
                                except Teacher.DoesNotExist:
                                    pass
                        elif attendance.classroom and attendance.classroom.class_teacher:
                            teacher = attendance.classroom.class_teacher
                            if teacher and teacher.user:
                                teacher_user = teacher.user
                        
                        if teacher_user:
                            coordinator_name = coordinator.full_name if coordinator else request.user.get_full_name() or request.user.username
                            classroom_name = str(attendance.classroom)
                            verb = "Your attendance has been approved"
                            target_text = f"by {coordinator_name} for {classroom_name} on {attendance.date.strftime('%B %d, %Y')}."
                            
                            from notifications.services import create_notification
                            create_notification(
                                recipient=teacher_user,
                                actor=request.user,
                                verb=verb,
                                target_text=target_text,
                                data={"attendance_id": attendance.id, "classroom_id": attendance.classroom.id}
                            )
                    except Exception as e:
                        # Don't fail bulk approval if notification fails
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Failed to send notification for attendance {attendance_id}: {str(e)}")
                    
                    approved_count += 1
                    
                except Exception as e:
                    failed_count += 1
                    errors.append(f"Attendance {attendance_id}: {str(e)}")
        
        return Response({
            'approved_count': approved_count,
            'failed_count': failed_count,
            'total': len(attendance_ids),
            'errors': errors[:10]  # Limit errors to first 10
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reopen_attendance(request, attendance_id):
    """Coordinator reopens finalized attendance with reason"""
    try:
        attendance = get_object_or_404(Attendance, id=attendance_id, is_deleted=False)
        reason = request.data.get('reason')
        
        if not reason:
            return Response({'error': 'Reason is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        if attendance.status != 'approved':
            return Response({'error': 'Can only reopen approved attendance'}, status=status.HTTP_400_BAD_REQUEST)
        
        from coordinator.models import Coordinator
        coordinator = Coordinator.get_for_user(request.user)
        if not coordinator:
            return Response({'error': 'Coordinator profile not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check membership - always check assigned_levels first, fallback to level field
        allowed = False
        if coordinator.assigned_levels.exists():
            if attendance.classroom.grade.level in coordinator.assigned_levels.all():
                allowed = True
        elif coordinator.level:
            if coordinator.level == attendance.classroom.grade.level:
                allowed = True

        if not allowed:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        with transaction.atomic():
            attendance.status = 'under_review'
            attendance.is_final = False
            attendance.reopened_at = timezone.now()
            attendance.reopened_by = request.user
            attendance.reopen_reason = reason
            attendance.add_edit_history(request.user, 'reopen', reason)
            attendance.save()
            
            from .models import AuditLog
            AuditLog.objects.create(
                feature='attendance',
                action='reopen',
                entity_type='Attendance',
                entity_id=attendance.id,
                user=request.user,
                ip_address=request.META.get('REMOTE_ADDR'),
                changes={'status': 'under_review'},
                reason=reason
            )
        
        return Response({'message': 'Attendance reopened successfully'})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def grant_backfill_permission(request):
    """Coordinator grants permission to mark attendance for missed date"""
    try:
        classroom_id = request.data.get('classroom_id')
        date_str = request.data.get('date')
        teacher_id = request.data.get('teacher_id')
        reason = request.data.get('reason')
        deadline_str = request.data.get('deadline')
        
        if not all([classroom_id, date_str, teacher_id, reason, deadline_str]):
            return Response({'error': 'All fields required'}, status=status.HTTP_400_BAD_REQUEST)
        
        classroom = get_object_or_404(ClassRoom, id=classroom_id)
        teacher = get_object_or_404(User, id=teacher_id)
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        deadline = datetime.strptime(deadline_str, '%Y-%m-%dT%H:%M:%S')
        
        from coordinator.models import Coordinator
        coordinator = Coordinator.get_for_user(request.user)
        if not coordinator:
            return Response({'error': 'Coordinator profile not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check membership for backfill permission
        allowed = False
        if coordinator.shift == 'both':
            if coordinator.assigned_levels.exists() and classroom.grade.level in coordinator.assigned_levels.all():
                allowed = True
            elif coordinator.level and coordinator.level == classroom.grade.level:
                allowed = True
        else:
            if coordinator.level and coordinator.level == classroom.grade.level:
                allowed = True

        if not allowed:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        from .models import AttendanceBackfillPermission, AuditLog
        permission = AttendanceBackfillPermission.objects.create(
            classroom=classroom,
            date=date_obj,
            granted_to=teacher,
            granted_by=request.user,
            reason=reason,
            deadline=deadline
        )
        
        AuditLog.objects.create(
            feature='attendance',
            action='approve',
            entity_type='AttendanceBackfillPermission',
            entity_id=permission.id,
            user=request.user,
            ip_address=request.META.get('REMOTE_ADDR'),
            reason=reason
        )
        
        return Response({'message': 'Backfill permission granted', 'permission_id': permission.id})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_backfill_permissions(request):
    """Get active backfill permissions for current user"""
    try:
        from .models import AttendanceBackfillPermission
        permissions = AttendanceBackfillPermission.objects.filter(
            granted_to=request.user,
            is_used=False
        ).select_related('classroom', 'granted_by')
        
        data = [{
            'id': p.id,
            'classroom_id': p.classroom.id,
            'classroom_name': str(p.classroom),
            'date': p.date,
            'reason': p.reason,
            'deadline': p.deadline,
            'is_expired': p.is_expired,
            'granted_by': p.granted_by.get_full_name() if p.granted_by else None
        } for p in permissions]
        
        return Response(data)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_holiday(request):
    """Coordinator creates holiday for their level(s) with optional grade selection"""
    try:
        date_str = request.data.get('date')
        reason = request.data.get('reason')
        # Support both old single level_id and new level_ids array
        level_id = request.data.get('level_id')  # Backward compatibility
        level_ids = request.data.get('level_ids', [])  # New: multiple levels
        grade_ids = request.data.get('grade_ids', [])  # New: optional grades
        shift_value = request.data.get('shift')
        allowed_shifts = resolve_allowed_shifts(shift_value)

        if not all([date_str, reason]):
            return Response({'error': 'Date and reason required'}, status=status.HTTP_400_BAD_REQUEST)

        from coordinator.models import Coordinator
        from classes.models import Level, Grade
        coordinator = Coordinator.get_for_user(request.user)
        if not coordinator:
            return Response({'error': 'Coordinator profile not found'}, status=status.HTTP_404_NOT_FOUND)

        # Determine which levels to use
        target_levels = []

        # Handle backward compatibility: if level_id provided, convert to level_ids
        if level_id and not level_ids:
            level_ids = [level_id]

        if level_ids:
            # Coordinator selected specific level(s)
            for lid in level_ids:
                target_level = get_object_or_404(Level, id=lid)
                # Verify coordinator has access to this level
                allowed = False
                if coordinator.assigned_levels.exists():
                    if target_level in coordinator.assigned_levels.all():
                        allowed = True
                elif coordinator.level == target_level:
                    allowed = True

                if not allowed and not request.user.is_superuser:
                    return Response({'error': f'Access denied to level {target_level.name}'}, status=status.HTTP_403_FORBIDDEN)

                target_levels.append(target_level)
        else:
            # Use coordinator's default level(s)
            if coordinator.assigned_levels.exists():
                target_levels = list(coordinator.assigned_levels.all())
            elif coordinator.level:
                target_levels = [coordinator.level]
            else:
                return Response({'error': 'No level assigned to coordinator'}, status=status.HTTP_400_BAD_REQUEST)

        if not target_levels:
            return Response({'error': 'At least one level must be selected'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            validate_levels_for_shift(target_levels, allowed_shifts)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()

        from .models import Holiday, AuditLog
        from classes.models import ClassRoom
        from django.utils import timezone

        # Validate grades if provided
        target_grades = []
        if grade_ids:
            for gid in grade_ids:
                grade = get_object_or_404(Grade, id=gid)
                # Verify grade belongs to one of the selected levels
                if grade.level not in target_levels:
                    return Response({'error': f'Grade {grade.name} does not belong to selected levels'}, status=status.HTTP_400_BAD_REQUEST)
                target_grades.append(grade)
        try:
            validate_grades_for_levels(target_grades, target_levels)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        # Check if date is in the past and archive existing attendance
        if date_obj < timezone.now().date():
            # Determine which classrooms to archive
            if target_grades:
                # Archive only classrooms in selected grades
                classrooms = ClassRoom.objects.filter(grade__in=target_grades)
            else:
                # Archive all classrooms in selected levels
                classrooms = ClassRoom.objects.filter(grade__level__in=target_levels)

            for classroom in classrooms:
                try:
                    existing_attendance = Attendance.objects.get(
                        classroom=classroom,
                        date=date_obj,
                        is_deleted=False
                    )

                    # Archive the attendance data
                    student_attendance_list = []
                    for sa in existing_attendance.student_attendances.all():
                        student_attendance_list.append({
                            'id': sa.id,
                            'student_id': sa.student_id,
                            'status': sa.status,
                            'remarks': sa.remarks or '',
                            'created_at': sa.created_at.isoformat() if sa.created_at else None,
                            'updated_at': sa.updated_at.isoformat() if sa.updated_at else None,
                        })

                    archived_data = {
                        'student_attendance': student_attendance_list,
                        'marked_by': existing_attendance.marked_by.get_full_name() if existing_attendance.marked_by else None,
                        'marked_at': existing_attendance.marked_at.isoformat(),
                        'status': existing_attendance.status,
                        'total_students': existing_attendance.total_students,
                        'present_count': existing_attendance.present_count,
                        'absent_count': existing_attendance.absent_count,
                        'late_count': existing_attendance.late_count,
                        'leave_count': existing_attendance.leave_count
                    }

                    # Mark as replaced by holiday
                    existing_attendance.replaced_by_holiday = True
                    existing_attendance.replaced_at = timezone.now()
                    existing_attendance.archived_data = archived_data
                    existing_attendance.save()

                except Attendance.DoesNotExist:
                    pass

        # Create holiday (use first level for backward compatibility in level field)
        holiday = Holiday.objects.create(
            date=date_obj,
            reason=reason,
            level=target_levels[0] if target_levels else None,
            created_by=request.user,
            organization=request.user.organization
        )

        # Assign relationships
        holiday.levels.set(target_levels)
        if target_grades:
            holiday.grades.set(target_grades)
        holiday.shifts = sorted(collect_shifts_from_levels(target_levels))
        holiday.save()

        AuditLog.objects.create(
            feature='attendance',
            action='create',
            entity_type='Holiday',
            entity_id=holiday.id,
            user=request.user,
            ip_address=request.META.get('REMOTE_ADDR'),
            reason=reason
        )

        # Notifications are handled centrally via signals
        return Response({'message': 'Holiday created', 'holiday_id': holiday.id})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_holidays(request):
    """Get holidays for user's level(s) with optional grade filtering"""
    try:
        level_id = request.query_params.get('level_id')
        level_ids = request.query_params.getlist('level_ids')  # Support multiple level_ids
        grade_id = request.query_params.get('grade_id')
        grade_ids = request.query_params.getlist('grade_ids')  # Support multiple grade_ids
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        shift_param = request.query_params.get('shift')
        allowed_shifts = resolve_allowed_shifts(shift_param)
        
        from .models import Holiday
        from classes.models import Level
        
        # Build query - support both old (level) and new (levels) fields
        holidays = Holiday.objects.all()
        
        # Filter by level(s)
        if level_ids:
            # New: filter by multiple levels using levels M2M
            holidays = holidays.filter(
                Q(levels__id__in=level_ids) | Q(level_id__in=level_ids)
            ).distinct()
        elif level_id:
            # Backward compatibility: single level
            holidays = holidays.filter(
                Q(levels__id=level_id) | Q(level_id=level_id)
            ).distinct()
        
        if allowed_shifts:
            holidays = holidays.filter(
                Q(levels__shift__in=allowed_shifts) |
                Q(level__shift__in=allowed_shifts) |
                Q(shifts__contains=sorted(allowed_shifts))
            ).distinct()

        # Filter by grade(s) if provided
        if grade_ids:
            holidays = holidays.filter(grades__id__in=grade_ids).distinct()
        elif grade_id:
            holidays = holidays.filter(grades__id=grade_id).distinct()
        
        if start_date:
            holidays = holidays.filter(date__gte=start_date)
        if end_date:
            holidays = holidays.filter(date__lte=end_date)
        
        # Serialize holidays with levels and grades
        data = []
        for h in holidays:
            # Get all levels (from levels M2M or fallback to level field)
            holiday_levels = list(h.levels.all())
            if not holiday_levels and h.level:
                holiday_levels = [h.level]
            
            # Get all grades
            holiday_grades = list(h.grades.all())

            holiday_shifts = h.shifts or sorted(collect_shifts_from_levels(holiday_levels))
            
            holiday_data = {
                'id': h.id,
                'date': h.date.strftime('%Y-%m-%d'),
                'reason': h.reason,
                'level_id': holiday_levels[0].id if holiday_levels else None,  # Backward compatibility
                'level_name': str(holiday_levels[0]) if holiday_levels else None,
                'level_ids': [l.id for l in holiday_levels],
                'level_names': [str(l) for l in holiday_levels],
                'grade_ids': [g.id for g in holiday_grades],
                'grade_names': [g.name for g in holiday_grades],
                'shifts': holiday_shifts,
                'created_by': h.created_by.get_full_name() if h.created_by else None
            }
            data.append(holiday_data)
        
        return Response(data)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_holiday(request, holiday_id):
    """Update existing holiday"""
    try:
        from .models import Holiday, AuditLog
        from coordinator.models import Coordinator
        from classes.models import ClassRoom
        from django.utils import timezone
        
        holiday = get_object_or_404(Holiday, id=holiday_id)
        
        # Verify coordinator has access to this holiday's level(s)
        coordinator = Coordinator.get_for_user(request.user)
        if not coordinator:
            return Response({'error': 'Coordinator profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get holiday levels (from levels M2M or fallback to level field)
        holiday_levels = list(holiday.levels.all())
        if not holiday_levels and holiday.level:
            holiday_levels = [holiday.level]
        
        # Check if coordinator manages any of the holiday's levels
        allowed = False
        if coordinator.assigned_levels.exists():
            if any(level in coordinator.assigned_levels.all() for level in holiday_levels):
                allowed = True
        elif coordinator.level in holiday_levels:
            allowed = True
        
        if not allowed and not request.user.is_superuser:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Check if holiday can be edited (must be at least 12 hours before holiday date)
        holiday_date = holiday.date
        # Create datetime at start of holiday date (midnight)
        holiday_datetime = timezone.make_aware(datetime.combine(holiday_date, datetime.min.time()))
        # Calculate 12 hours before holiday date
        twelve_hours_before = holiday_datetime - timedelta(hours=12)
        
        if timezone.now() >= twelve_hours_before:
            return Response({
                'error': f'Cannot edit holiday within 12 hours of the holiday date. Holiday is on {holiday_date.strftime("%B %d, %Y")}.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        date_str = request.data.get('date')
        reason = request.data.get('reason')
        # Support both old single level_id and new level_ids array
        level_id = request.data.get('level_id')  # Backward compatibility
        level_ids = request.data.get('level_ids', [])  # New: multiple levels
        grade_ids = request.data.get('grade_ids', [])  # New: optional grades
        shift_value = request.data.get('shift')
        allowed_shifts = resolve_allowed_shifts(shift_value)
        
        if not all([date_str, reason]):
            return Response({'error': 'Date and reason required'}, status=status.HTTP_400_BAD_REQUEST)
        
        from classes.models import Level, Grade
        
        # Handle backward compatibility: if level_id provided, convert to level_ids
        if level_id and not level_ids:
            level_ids = [level_id]
        
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        old_date = holiday.date
        old_levels = list(holiday.levels.all())
        if not old_levels and holiday.level:
            old_levels = [holiday.level]
        old_grades = list(holiday.grades.all())
        
        # Determine new levels
        new_levels = []
        if level_ids:
            for lid in level_ids:
                level = get_object_or_404(Level, id=lid)
                # Verify coordinator has access
                allowed = False
                if coordinator.assigned_levels.exists():
                    if level in coordinator.assigned_levels.all():
                        allowed = True
                elif coordinator.level == level:
                    allowed = True
                
                if not allowed and not request.user.is_superuser:
                    return Response({'error': f'Access denied to level {level.name}'}, status=status.HTTP_403_FORBIDDEN)
                
                new_levels.append(level)
        else:
            # Keep existing levels
            new_levels = old_levels

        try:
            validate_levels_for_shift(new_levels, allowed_shifts)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate grades if provided
        new_grades = []
        if grade_ids:
            for gid in grade_ids:
                grade = get_object_or_404(Grade, id=gid)
                # Verify grade belongs to one of the selected levels
                if grade.level not in new_levels:
                    return Response({'error': f'Grade {grade.name} does not belong to selected levels'}, status=status.HTTP_400_BAD_REQUEST)
                new_grades.append(grade)
        try:
            validate_grades_for_levels(new_grades, new_levels)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        
        # If date or levels/grades changed, handle existing attendance
        levels_changed = set([l.id for l in old_levels]) != set([l.id for l in new_levels])
        grades_changed = set([g.id for g in old_grades]) != set([g.id for g in new_grades])
        
        if date_obj != old_date or levels_changed or grades_changed:
            # Archive attendance for old date/levels/grades if in past
            if old_date < timezone.now().date():
                # Determine which classrooms to restore
                if old_grades:
                    # Restore only classrooms in old grades
                    classrooms = ClassRoom.objects.filter(grade__in=old_grades)
                else:
                    # Restore all classrooms in old levels
                    classrooms = ClassRoom.objects.filter(grade__level__in=old_levels)
                
                for classroom in classrooms:
                    try:
                        existing_attendance = Attendance.objects.get(
                            classroom=classroom,
                            date=old_date,
                            is_deleted=False,
                            replaced_by_holiday=True
                        )
                        # Restore archived attendance if it exists
                        if existing_attendance.archived_data:
                            existing_attendance.replaced_by_holiday = False
                            existing_attendance.save()
                    except Attendance.DoesNotExist:
                        pass
            
            # Archive attendance for new date/levels/grades if in past
            if date_obj < timezone.now().date():
                # Determine which classrooms to archive
                if new_grades:
                    # Archive only classrooms in new grades
                    classrooms = ClassRoom.objects.filter(grade__in=new_grades)
                else:
                    # Archive all classrooms in new levels
                    classrooms = ClassRoom.objects.filter(grade__level__in=new_levels)
                
                for classroom in classrooms:
                    try:
                        existing_attendance = Attendance.objects.get(
                            classroom=classroom,
                            date=date_obj,
                            is_deleted=False
                        )
                        # Archive the attendance data
                        student_attendance_list = []
                        for sa in existing_attendance.student_attendances.all():
                            student_attendance_list.append({
                                'id': sa.id,
                                'student_id': sa.student_id,
                                'status': sa.status,
                                'remarks': sa.remarks or '',
                                'created_at': sa.created_at.isoformat() if sa.created_at else None,
                                'updated_at': sa.updated_at.isoformat() if sa.updated_at else None,
                            })
                        
                        archived_data = {
                            'student_attendance': student_attendance_list,
                            'marked_by': existing_attendance.marked_by.get_full_name() if existing_attendance.marked_by else None,
                            'marked_at': existing_attendance.marked_at.isoformat(),
                            'status': existing_attendance.status,
                            'total_students': existing_attendance.total_students,
                            'present_count': existing_attendance.present_count,
                            'absent_count': existing_attendance.absent_count,
                            'late_count': existing_attendance.late_count,
                            'leave_count': existing_attendance.leave_count
                        }
                        existing_attendance.replaced_by_holiday = True
                        existing_attendance.replaced_at = timezone.now()
                        existing_attendance.archived_data = archived_data
                        existing_attendance.save()
                    except Attendance.DoesNotExist:
                        pass
        
        # Update holiday
        holiday.date = date_obj
        holiday.reason = reason
        # Update level for backward compatibility (use first level)
        holiday.level = new_levels[0] if new_levels else None
        if not holiday.organization and request.user.organization:
            holiday.organization = request.user.organization
        holiday.save()
        
        # Update levels M2M
        holiday.levels.set(new_levels)
        
        # Update grades M2M
        holiday.grades.set(new_grades)

        # Store resolved shifts
        holiday.shifts = sorted(collect_shifts_from_levels(new_levels))
        
        holiday.save()
        
        AuditLog.objects.create(
            feature='attendance',
            action='update',
            entity_type='Holiday',
            entity_id=holiday.id,
            user=request.user,
            ip_address=request.META.get('REMOTE_ADDR'),
            reason=reason
        )
        
        return Response({'message': 'Holiday updated successfully', 'holiday_id': holiday.id})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_holiday(request, holiday_id):
    """Delete holiday and optionally restore archived attendance"""
    try:
        from .models import Holiday, AuditLog
        from coordinator.models import Coordinator
        from classes.models import ClassRoom
        from django.utils import timezone

        holiday = get_object_or_404(Holiday, id=holiday_id)

        # Verify coordinator has access
        coordinator = Coordinator.get_for_user(request.user)
        if not coordinator:
            return Response({'error': 'Coordinator profile not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check if coordinator manages this holiday
        holiday_levels = list(holiday.levels.all())
        if not holiday_levels and holiday.level:
            holiday_levels = [holiday.level]

        allowed = False
        if coordinator.shift == 'both':
            if coordinator.assigned_levels.exists():
                if any(level in coordinator.assigned_levels.all() for level in holiday_levels):
                    allowed = True
            elif holiday.level and coordinator.level == holiday.level:
                allowed = True
        else:
            if holiday.level and coordinator.level == holiday.level:
                allowed = True

        if not allowed and not request.user.is_superuser:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)

        # Check if holiday can be deleted (must be at least 12 hours before holiday date)
        holiday_date = holiday.date
        holiday_datetime = timezone.make_aware(datetime.combine(holiday_date, datetime.min.time()))
        twelve_hours_before = holiday_datetime - timedelta(hours=12)
        if timezone.now() >= twelve_hours_before:
            return Response({'error': 'Cannot delete holiday within 12 hours of the holiday date.'}, status=status.HTTP_403_FORBIDDEN)

        restore_attendance = request.data.get('restore_attendance', False)

        # Restore archived attendance if requested
        if restore_attendance:
            classrooms = ClassRoom.objects.filter(grade__level__in=holiday_levels)
            for classroom in classrooms:
                try:
                    attendance = Attendance.objects.get(
                        classroom=classroom,
                        date=holiday_date,
                        is_deleted=False,
                        replaced_by_holiday=True
                    )
                    attendance.replaced_by_holiday = False
                    attendance.replaced_at = None
                    attendance.save()
                except Attendance.DoesNotExist:
                    continue

        holiday.delete()

        AuditLog.objects.create(
            feature='attendance',
            action='delete',
            entity_type='Holiday',
            entity_id=holiday_id,
            user=request.user,
            ip_address=request.META.get('REMOTE_ADDR'),
            reason=f'Deleted holiday: {holiday.reason}'
        )

        # Notifications handled via signals
        return Response({'message': 'Holiday deleted successfully'})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_realtime_attendance_metrics(request):
    """Get real-time attendance metrics for dashboards"""
    try:
        user = request.user
        today = timezone.now().date()
        
        metrics = {
            'today': today.isoformat(),
            'classrooms': []
        }
        
        # Get classrooms based on role
        if user.is_teacher():
            teacher = Teacher.objects.get(employee_code=user.username)
            classrooms = [teacher.assigned_classroom] if teacher.assigned_classroom else []
        elif user.is_coordinator():
            from coordinator.models import Coordinator
            coordinator = Coordinator.get_for_user(user)
            if coordinator:
                if coordinator.assigned_levels.exists():
                    classrooms = ClassRoom.objects.filter(grade__level__in=coordinator.assigned_levels.all())
                elif coordinator.level:
                    classrooms = ClassRoom.objects.filter(grade__level=coordinator.level)
                else:
                    classrooms = []
            else:
                classrooms = []
        elif user.is_principal():
            from principals.models import Principal
            principal = Principal.objects.get(email=user.email)
            classrooms = ClassRoom.objects.filter(grade__level__campus=principal.campus)
        else:
            classrooms = []
        
        for classroom in classrooms:
            attendance = Attendance.objects.filter(
                classroom=classroom,
                date=today
            ).first()
            
            status_color = 'gray'
            if attendance:
                if attendance.status == 'draft':
                    status_color = 'yellow'
                elif attendance.status == 'submitted':
                    status_color = 'blue'
                elif attendance.status == 'under_review':
                    status_color = 'orange'
                elif attendance.status == 'approved':
                    status_color = 'green'
            
            metrics['classrooms'].append({
                'id': classroom.id,
                'name': str(classroom),
                'status': attendance.status if attendance else 'not_marked',
                'status_color': status_color,
                'total_students': attendance.total_students if attendance else classroom.students.count(),
                'present_count': attendance.present_count if attendance else 0,
                'absent_count': attendance.absent_count if attendance else 0,
                'percentage': attendance.attendance_percentage if attendance else 0
            })
        
        return Response(metrics)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_attendance_list(request):
    """
    Get list of attendance records for dashboard
    """
    try:
        # Optional campus filter: restrict attendance to classrooms
        # whose grade.level.campus matches the given campus ID.
        campus_id = request.GET.get('campus')

        # For principals, default to their campus if no explicit campus param is provided
        user = request.user
        if not campus_id and hasattr(user, 'campus') and user.campus:
            campus_id = str(user.campus.id)

        # Accept configurable days param (default 30, max 180)
        try:
            days = min(int(request.GET.get('days', 30)), 180)
        except (ValueError, TypeError):
            days = 30
        since = timezone.now().date() - timedelta(days=days)

        attendances = Attendance.objects.filter(
            date__gte=since,
            is_deleted=False
        ).select_related('classroom')

        # Role-based Data Isolation
        if not user.is_superadmin():
            if user.role == 'admin':
                # Reseller Admin: See all organizations created by them
                attendances = attendances.filter(organization__created_by=user)
            elif user.role == 'org_admin' and user.organization:
                # Org Admin: Strictly filter by their assigned organization
                attendances = attendances.filter(organization=user.organization)
            elif user.organization:
                # Other staff: Filter by organization
                attendances = attendances.filter(organization=user.organization)
            else:
                # If no organization and not superadmin, they shouldn't see any data
                return Response([], status=status.HTTP_200_OK)

        if campus_id:
            try:
                campus_id_int = int(campus_id)
                attendances = attendances.filter(
                    classroom__grade__level__campus_id=campus_id_int
                )
            except ValueError:
                # Ignore invalid campus values and fall back to all campuses
                pass

        attendances = attendances.order_by('-date')

        # Return only summary fields needed by the dashboard chart (avoids heavy nested serialization)
        data = list(attendances.values('date', 'present_count', 'absent_count', 'total_students'))
        return Response(data, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_delete_logs(request):
    """
    Get delete audit logs for the current user or all (based on permissions)
    """
    try:
        from .models import AuditLog
        from .serializers import DeleteLogSerializer
        
        # Get query parameters
        feature = request.GET.get('feature')  # Optional filter by feature
        limit = request.GET.get('limit', 50)  # Default 50, max 200
        try:
            limit = min(int(limit), 200)
        except (ValueError, TypeError):
            limit = 50
        
        # Base queryset - only delete actions
        queryset = AuditLog.objects.filter(action='delete').select_related('user').order_by('-timestamp')
        
        # Filter by feature if provided
        if feature:
            queryset = queryset.filter(feature=feature)
        
        # Permission-based filtering
        user = request.user
        if user.is_superuser or (hasattr(user, 'role') and user.role in ['superadmin', 'org_admin']):
            # Super admin and Org admin can see delete logs
            # For Org admin, they are already filtered by OrganizationManager
            pass
        elif hasattr(user, 'role') and user.role == 'principal':
            # Principal can see delete logs for their campus
            if user.campus:
                # For student deletions, check if student belongs to principal's campus
                # Get all student IDs in principal's campus (including soft-deleted)
                campus_student_ids = Student.objects.with_deleted().filter(
                    campus=user.campus
                ).values_list('id', flat=True)
                
                # Filter delete logs: either feature is campus-related OR student_id is in campus
                queryset = queryset.filter(
                    Q(feature__in=['teacher', 'coordinator', 'classroom', 'grade', 'level']) |
                    Q(feature='student', entity_id__in=campus_student_ids) |
                    Q(changes__campus_id=user.campus.id) |
                    Q(changes__campus=user.campus.id)
                )
        elif hasattr(user, 'role') and user.role == 'coordinator':
            # Coordinator can see delete logs for students in their managed classrooms
            try:
                coordinator_obj = Coordinator.get_for_user(user)
                if coordinator_obj and coordinator_obj.campus:
                    # Determine which levels this coordinator manages
                    managed_levels = []
                    if coordinator_obj.assigned_levels.exists():
                        managed_levels = list(coordinator_obj.assigned_levels.all())
                    if not managed_levels and coordinator_obj.level:
                        managed_levels = [coordinator_obj.level]

                    if managed_levels:
                        # Get all classrooms under coordinator's managed levels
                        coordinator_classrooms = ClassRoom.objects.filter(
                            grade__level__in=managed_levels,
                            grade__level__campus=coordinator_obj.campus
                        ).values_list('id', flat=True)

                        # Get all student IDs in these classrooms (including soft-deleted)
                        coordinator_student_ids = Student.objects.with_deleted().filter(
                            classroom__in=coordinator_classrooms
                        ).values_list('id', flat=True)

                        # Filter delete logs: show student deletions for students in coordinator's classrooms
                        # Also show other relevant features (teacher, classroom, etc.) if they relate to coordinator's scope
                        queryset = queryset.filter(
                            Q(feature='student', entity_id__in=coordinator_student_ids) |
                            Q(feature__in=['teacher', 'classroom', 'grade', 'level'])
                        )
                    else:
                        # No managed levels, show only their own delete logs
                        queryset = queryset.filter(user=user)
                else:
                    # No coordinator profile found or no campus, show only their own delete logs
                    queryset = queryset.filter(user=user)
            except Exception:
                # If coordinator resolution fails, show only their own delete logs
                queryset = queryset.filter(user=user)
        else:
            # Other users (teachers, etc.) see only their own delete logs
            queryset = queryset.filter(user=user)
        
        # Limit results
        delete_logs = queryset[:limit]
        
        # Serialize the data
        serializer = DeleteLogSerializer(delete_logs, many=True)
        
        return Response({
            'results': serializer.data,
            'count': len(serializer.data),
            'total': queryset.count()
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_monthly_attendance(request, student_id):
    """Calculate attendance summary for a student in a specific month"""
    import calendar as cal_module
    from .models import Holiday

    student = get_object_or_404(Student, id=student_id)
    month_name = request.GET.get('month')
    academic_year = request.GET.get('year', '2024-25')

    # Academic year "2025-26": Apr-Dec → 2025, Jan-Mar → 2026
    calendar_year = int(academic_year.split('-')[0])

    month_map = {
        'April': 4, 'May': 5, 'June': 6, 'July': 7, 'August': 8, 'September': 9,
        'October': 10, 'November': 11, 'December': 12,
        'January': 1, 'February': 2, 'March': 3,
    }

    month_num = month_map.get(month_name)
    if not month_num:
        return Response({'error': 'Invalid month name'}, status=status.HTTP_400_BAD_REQUEST)

    if month_num <= 3:  # Jan–Mar belong to the second calendar year of the academic year
        calendar_year += 1

    classroom = student.classroom
    if not classroom:
        return Response({'days_present': 0, 'total_working_days': 0})

    # ── Step 1: total calendar days in this month ──────────────────────────────
    total_days_in_month = cal_module.monthrange(calendar_year, month_num)[1]

    # ── Step 2: count Sundays (weekday 6) in this month ───────────────────────
    sundays = sum(
        1 for day in range(1, total_days_in_month + 1)
        if date(calendar_year, month_num, day).weekday() == 6
    )

    # ── Step 3: count holidays for this level/org in this month ───────────────
    level = None
    if hasattr(classroom, 'grade') and classroom.grade and hasattr(classroom.grade, 'level'):
        level = classroom.grade.level

    holiday_qs = Holiday.objects.filter(
        date__year=calendar_year,
        date__month=month_num,
        organization=student.organization,
    )
    if level:
        holiday_qs = holiday_qs.filter(Q(levels=level) | Q(level=level))

    # Don't double-count holidays that already fall on Sunday
    holiday_dates = set(holiday_qs.values_list('date', flat=True))
    non_sunday_holidays = sum(1 for d in holiday_dates if d.weekday() != 6)

    # Count how many attendance records were actually marked for this classroom in this month
    actual_marked_days = Attendance.objects.filter(
        classroom=classroom,
        date__year=calendar_year,
        date__month=month_num,
        is_deleted=False
    ).count()

    calendar_working_days = total_days_in_month - sundays - non_sunday_holidays
    total_working_days = actual_marked_days if actual_marked_days > 0 else calendar_working_days

    # ── Days present (present + late count; leave does NOT count as present) ──
    days_present = StudentAttendance.objects.filter(
        student=student,
        attendance__date__year=calendar_year,
        attendance__date__month=month_num,
        attendance__is_deleted=False,
        status__in=['present', 'late'],
    ).count()

    return Response({
        'days_present': days_present,
        'total_working_days': total_working_days,
        'calendar_working_days': calendar_working_days,
    })



# ─── ZKTeco Device Management ────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def zkteco_devices(request):
    org = getattr(request.user, 'organization', None)

    if request.method == 'GET':
        devices = ZKTecoDevice.objects.filter(organization=org)
        serializer = ZKTecoDeviceSerializer(devices, many=True)
        return Response(serializer.data)

    if request.method == 'POST':
        role = getattr(request.user, 'role', '')
        if role not in ('org_admin', 'superadmin', 'principal'):
            return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)
        serializer = ZKTecoDeviceSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(organization=org)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def zkteco_device_detail(request, device_id):
    org = getattr(request.user, 'organization', None)
    device = get_object_or_404(ZKTecoDevice, id=device_id, organization=org)

    if request.method == 'GET':
        return Response(ZKTecoDeviceSerializer(device).data)

    role = getattr(request.user, 'role', '')
    if role not in ('org_admin', 'superadmin', 'principal'):
        return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'PUT':
        serializer = ZKTecoDeviceSerializer(device, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'DELETE':
        device.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ─── ZKTeco Employee Mappings ─────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def zkteco_mappings(request):
    org = getattr(request.user, 'organization', None)

    if request.method == 'GET':
        device_id = request.query_params.get('device_id')
        qs = ZKTecoEmployeeMapping.objects.filter(organization=org).select_related('device', 'user', 'teacher')
        if device_id:
            qs = qs.filter(device_id=device_id)
        serializer = ZKTecoMappingSerializer(qs, many=True)
        return Response(serializer.data)

    if request.method == 'POST':
        role = getattr(request.user, 'role', '')
        if role not in ('org_admin', 'superadmin', 'principal'):
            return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        # Allow bulk create: accept list or single object
        many = isinstance(request.data, list)
        serializer = ZKTecoMappingSerializer(data=request.data, many=many)
        if serializer.is_valid():
            if many:
                serializer.save(organization=org, created_by=request.user)
            else:
                serializer.save(organization=org, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def zkteco_mapping_detail(request, mapping_id):
    org = getattr(request.user, 'organization', None)
    mapping = get_object_or_404(ZKTecoEmployeeMapping, id=mapping_id, organization=org)

    role = getattr(request.user, 'role', '')
    if role not in ('org_admin', 'superadmin', 'principal'):
        return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'PUT':
        serializer = ZKTecoMappingSerializer(mapping, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'DELETE':
        mapping.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def zkteco_unmapped_staff(request):
    """Staff members (Users) jo abhi kisi device se link nahi hain."""
    from users.models import User
    org = getattr(request.user, 'organization', None)
    device_id = request.query_params.get('device_id')

    mapped_user_ids = ZKTecoEmployeeMapping.objects.filter(
        organization=org, is_active=True, user__isnull=False
    )
    if device_id:
        mapped_user_ids = mapped_user_ids.filter(device_id=device_id)
    mapped_user_ids = mapped_user_ids.values_list('user_id', flat=True)

    staff = User.objects.filter(organization=org, is_active=True).exclude(
        role__in=['student', 'superadmin', 'donor']
    ).exclude(
        id__in=mapped_user_ids
    )
    
    result = []
    for s in staff:
        result.append({
            'id': s.id,
            'full_name': s.get_full_name(),
            'employee_code': s.username,
            'current_campus__campus_name': s.campus.campus_name if s.campus else None
        })

    return Response(result)


# ─── ZKTeco Push Receiver ─────────────────────────────────────────────────────

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def zkteco_push(request):
    """
    ZKTeco machine yahan punch data bhejti hai.
    GET: Machine heartbeat / device info check
    POST: Attendance punch receive karna

    ZKTeco PUSH format:
    {
        "user_id": "TCH001",
        "timestamp": "2026-04-17 08:30:00",
        "punch": 0,          # 0=check-in, 1=check-out, 4=break-out, 5=break-in
        "device_serial": "JYM6243800319"
    }
    """
    print(f"\n--- [ZKTECO DEBUG] ---")
    print(f"Method: {request.method}")
    print(f"Path: {request.path}")
    print(f"Query Params: {request.GET}")
    print(f"Body (Raw): {request.body[:500]}")
    
    from django.conf import settings as django_settings

    # 1. Parse incoming data (DRF's request.data is not available in regular Django views)
    import json
    data = {}
    if request.content_type == 'application/json':
        try:
            data = json.loads(request.body)
        except:
            pass
    else:
        # Try to use request.POST if it's form-encoded
        data = request.POST.dict()

    # 1. Heartbeat & SN logic
    sn = request.GET.get('SN') or data.get('device_serial') or data.get('sn')
    device = None
    if sn:
        device = ZKTecoDevice.objects.filter(serial_number=sn, is_active=True).first()
        if device:
            device.last_sync = timezone.now()
            device.save(update_fields=['last_sync'])

    # Heartbeat response (GET)
    if request.method == 'GET':
        if sn:
            return HttpResponse("OK")
        return JsonResponse({
            'status': 'ok',
            'server_time': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
        })

    # 2. Parse Data (ADMS sends raw text with multiple lines)
    raw_body = request.body.decode('utf-8') if hasattr(request, 'body') else ""
    lines = raw_body.strip().split('\n')
    
    processed_count = 0
    from teachers.models import Teacher as TeacherModel
    from coordinator.models import Coordinator
    from principals.models import Principal
    from .models import StaffAttendance, ZKTecoEmployeeMapping

    for line in lines:
        line = line.strip()
        if not line: continue
        
        # Handle formats
        parts = line.split('\t')
        user_id = None
        timestamp_str = None
        punch_type = 0
        
        if len(parts) >= 2:
            if '=' not in parts[0]:
                user_id = parts[0].strip()
                timestamp_str = parts[1].strip()
                try:
                    punch_type = int(parts[2].strip()) if len(parts) > 2 else 0
                except: punch_type = 0
            else:
                data_map = {}
                for p in parts:
                    if '=' in p:
                        k, v = p.split('=', 1)
                        data_map[k.strip().upper()] = v.strip()
                user_id = data_map.get('USERID') or data_map.get('PIN')
                timestamp_str = data_map.get('TIMESTAMP') or data_map.get('TIME')
                try:
                    punch_type = int(data_map.get('STATUS') or 0)
                except: punch_type = 0

        if not user_id or not timestamp_str:
            print(f"--- [ZKTECO DEBUG] Invalid line: {line} ---")
            continue

        # 3. Timestamp parse
        try:
            timestamp_str = str(timestamp_str).strip()
            if 'T' in timestamp_str:
                punch_dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            else:
                punch_dt = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
            
            # Use machine local time
            punch_dt = timezone.make_aware(punch_dt) if timezone.is_naive(punch_dt) else punch_dt
            punch_date = punch_dt.date()
            punch_time = punch_dt.time()
        except Exception as e:
            print(f"--- [ZKTECO DEBUG] Timestamp error for user {user_id}: {e} ---")
            continue

        # 4. Teacher lookup
        teacher = None
        mapping = None
        # 4. Identify Staff (User)
        user_obj = None
        device_obj = ZKTecoDevice._base_manager.filter(serial_number=sn, is_active=True).first() if sn else None
        mapping  = ZKTecoEmployeeMapping._base_manager.filter(device=device_obj, device_user_id=str(user_id)).first() if device_obj else None
        
        if mapping and mapping.user:
            user_obj = mapping.user
        else:
            # Fallback: search User model by biometric_id or username
            uid = str(user_id)
            from users.models import User
            user_obj = User.objects.filter(organization=device_obj.organization, biometric_id=uid).first() or \
                       User.objects.filter(organization=device_obj.organization, username=uid).first()

        if not user_obj:
            print(f"--- [ZKTECO DEBUG] No staff/user found for ID: {user_id} ---")
            # Create mapping for unknown users for later linking
            if device_obj:
                ZKTecoEmployeeMapping._base_manager.get_or_create(
                    device=device_obj,
                    device_user_id=str(user_id),
                    defaults={
                        'organization': device_obj.organization,
                        'is_active': True, 
                        'device_user_name': f"User {user_id}"
                    }
                )
            continue

        # --- AUTO-MAPPING FOR FOUND STAFF ---
        if device_obj and not mapping:
            mapping, created = ZKTecoEmployeeMapping._base_manager.get_or_create(
                device=device_obj,
                device_user_id=str(user_id),
                defaults={
                    'organization': user_obj.organization,
                    'user': user_obj,
                    'is_active': True,
                    'device_user_name': user_obj.get_full_name()
                }
            )

        print(f"--- [ZKTECO DEBUG] Found Staff: {user_obj.get_full_name()} (Role: {user_obj.role}) ---")

        # 5. Timing logic (simplified for now — use default if no specific timing)
        # In a real scenario, we'd lookup EmployeeShiftTiming for this USER.
        shift_start = time(8, 0)
        grace = 10
        try:
            timing = user_obj.shift_timing
            if timing and timing.is_active:
                shift_start = timing.check_in_time
                grace = timing.grace_minutes
        except:
            pass

        # 6. Upsert StaffAttendance
        punch_type = int(punch_type)
        if punch_type in (0, 2, 3, 4):  # Check-in types
            rec, created = StaffAttendance._base_manager.get_or_create(
                user=user_obj,
                date=punch_date,
                defaults={
                    'organization': user_obj.organization,
                    'campus': user_obj.campus,
                    'check_in_time': punch_time,
                    'status': 'present',
                    'source': 'biometric',
                    'device': mapping.device if mapping else None,
                }
            )
            if created:
                late_mins = int((datetime.combine(punch_date, punch_time) - datetime.combine(punch_date, shift_start)).total_seconds() // 60)
                if late_mins > grace:
                    rec.late_minutes = late_mins
                    rec.status = 'late'
                    rec.save(update_fields=['late_minutes', 'status'])
            elif not rec.check_in_time or punch_time < rec.check_in_time:
                rec.check_in_time = punch_time
                rec.save(update_fields=['check_in_time'])
        
        elif punch_type == 1:  # Check-out
            rec, _ = StaffAttendance._base_manager.get_or_create(
                user=user_obj,
                date=punch_date,
                defaults={
                    'organization': user_obj.organization,
                    'campus': user_obj.campus,
                    'status': 'present',
                    'source': 'biometric',
                }
            )
            if not rec.check_out_time or punch_time > rec.check_out_time:
                rec.check_out_time = punch_time
                rec.save(update_fields=['check_out_time'])
        
        processed_count += 1
    
    print(f"--- [ZKTECO] Processed {processed_count} logs for SN: {sn} ---")
    return HttpResponse("OK")


def get_employee_timing(teacher):
    """Return EmployeeShiftTiming for a specific teacher if configured."""
    try:
        t = teacher.shift_timing
        return t if t.is_active else None
    except EmployeeShiftTiming.DoesNotExist:
        return None


def resolve_timing(teacher, org, campus, shift):
    """Employee timing. Hardcoded defaults fallback only."""
    emp = get_employee_timing(teacher)
    if emp:
        return emp.check_in_time, emp.check_out_time, emp.grace_minutes
    from datetime import time as time_type
    return time_type(8, 0), time_type(14, 0), 10


# ─── Teacher Attendance Views ─────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def staff_attendance_list(request):
    """
    GET /api/attendance/staff/?date=YYYY-MM-DD&campus_id=X
    Returns staff attendance for a given date (today by default).
    Universal visibility logic:
    - Coordinator: sees teachers at their campus.
    - Principal: sees teachers, coordinators, auditors, accountants, receptionists at their campus.
    - Org Admin: sees principals of their campuses.
    """
    from users.models import User
    from campus.models import Campus

    user = request.user
    org  = getattr(user, 'organization', None)

    date_str   = request.query_params.get('date', str(date.today()))
    campus_id  = request.query_params.get('campus_id')

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

    # Base staff QuerySet (exclude students and superadmins)
    staff_qs = User.objects.filter(organization=org, is_active=True).exclude(role__in=['student', 'superadmin', 'donor'])

    if user.is_principal():
        # Principal sees everyone at their campus except students and org admin
        staff_qs = staff_qs.filter(campus=user.campus).exclude(role='org_admin')
    elif user.is_coordinator():
        # Coordinator sees only teachers at their campus
        staff_qs = staff_qs.filter(campus=user.campus, role='teacher')
    elif user.is_org_admin_role():
        # Org Admin can filter by campus if provided
        if campus_id:
            staff_qs = staff_qs.filter(campus_id=campus_id)
        # By default (if no campus filtered), show principals (per requirement)
        else:
            staff_qs = staff_qs.filter(role='principal')
    
    # Fetch existing attendance records for that date
    existing = {
        rec.user_id: rec
        for rec in StaffAttendance.objects.filter(
            organization=org, date=target_date,
            user__in=staff_qs
        ).select_related('user', 'campus', 'marked_by')
    }

    result = []
    for staff in staff_qs:
        rec = existing.get(staff.id)
        if rec:
            data = StaffAttendanceSerializer(rec).data
        else:
            data = {
                'id': None,
                'user': staff.id,
                'staff_name': staff.get_full_name(),
                'staff_role': staff.get_role_display(),
                'staff_photo': staff.photo.url if staff.photo else None,
                'employee_code': staff.employee_code,
                'campus': staff.campus_id,
                'campus_name': staff.campus.campus_name if staff.campus else None,
                'date': str(target_date),
                'check_in_time': None,
                'check_out_time': None,
                'working_hours': None,
                'working_hours_display': None,
                'status': 'not_marked',
                'late_minutes': 0,
                'source': None,
                'device': None,
                'marked_by': None,
                'marked_by_name': None,
                'remarks': '',
                'created_at': None,
                'updated_at': None,
            }
        result.append(data)

    return Response(result)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def staff_attendance_mark(request):
    """
    Permission flow:
    - Coordinator  → can mark Teachers only
    - Principal    → can mark all staff at campus (including Coordinator)
    - Org Admin    → can mark Principals
    """
    from users.models import User

    user = request.user
    org  = getattr(user, 'organization', None)

    # Determine which staff_roles this user can mark
    if user.is_superuser or (hasattr(user, 'is_org_admin') and user.is_org_admin):
        allowed_roles = ['principal']
    elif user.is_principal():
        allowed_roles = ['teacher', 'coordinator', 'accounts_officer', 'admissions_counselor', 'compliance_officer']
    elif user.is_coordinator():
        allowed_roles = ['teacher']
    else:
        return Response({'error': 'You do not have permission to mark attendance.'}, status=status.HTTP_403_FORBIDDEN)

    date_str = request.data.get('date')
    records  = request.data.get('records', [])

    if not date_str or not records:
        return Response({'error': 'date and records are required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return Response({'error': 'Invalid date format.'}, status=status.HTTP_400_BAD_REQUEST)

    saved = []
    errors = []

    for item in records:
        try:
            target_user_id = item.get('staff_id') or item.get('user_id') or item.get('teacher_id')
            print(f"--- [ZKTECO DEBUG] Manual Mark Attempt for User ID: {target_user_id} ---")

            target_user = User.objects.get(id=target_user_id, organization=org)

            # Security: check if role is allowed
            if target_user.role not in allowed_roles:
                errors.append(f"Not allowed to mark {target_user.get_full_name()} (role {target_user.role})")
                continue

            # Security: check campus if not org admin
            if not user.is_superuser and not (hasattr(user, 'is_org_admin') and user.is_org_admin):
                if target_user.campus != user.campus:
                    errors.append(f"Not allowed to mark {target_user.get_full_name()} - different campus.")
                    continue

            # Correct logic: Principal can't mark themselves, Coordinator can't mark themselves.
            if target_user == user:
                errors.append(f"You cannot mark your own attendance.")
                continue

            # Upsert
            rec, created = StaffAttendance.objects.update_or_create(
                organization=org,
                user=target_user,
                date=target_date,
                defaults={
                    'campus': target_user.campus,
                    'status': item.get('status', 'absent'),
                    'check_in_time': item.get('check_in_time') or None,
                    'check_out_time': item.get('check_out_time') or None,
                    'remarks': item.get('remarks', ''),
                    'marked_by': user,
                    'source': 'manual'
                }
            )
            saved.append(StaffAttendanceSerializer(rec).data)
        except Exception as e:
            errors.append(str(e))

    # --- REMOVE OLD ATOMIC BLOCK (was part of original code) ---
    return Response({'saved': saved, 'errors': errors})
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def staff_attendance_summary(request):
    """
    Summary metrics for staff attendance.
    """
    from users.models import User
    user = request.user
    org  = getattr(user, 'organization', None)
    date_str = request.query_params.get('date', str(date.today()))
    campus_id = request.query_params.get('campus_id')

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return Response({'error': 'Invalid date.'}, status=status.HTTP_400_BAD_REQUEST)

    # Base staff QuerySet (filtering logic similar to list)
    staff_qs = User.objects.filter(organization=org, is_active=True).exclude(role__in=['student', 'superadmin', 'donor'])
    
    if user.is_principal():
        staff_qs = staff_qs.filter(campus=user.campus).exclude(role='org_admin')
    elif user.is_coordinator():
        staff_qs = staff_qs.filter(campus=user.campus, role='teacher')
    elif user.is_org_admin_role():
        if campus_id:
            staff_qs = staff_qs.filter(campus_id=campus_id)
        else:
            staff_qs = staff_qs.filter(role='principal')

    total = staff_qs.count()
    records = StaffAttendance.objects.filter(
        organization=org, date=target_date, user__in=staff_qs
    )

    counts = {s: 0 for s in ['present', 'absent', 'late', 'leave', 'half_day']}
    for r in records:
        if r.status in counts:
            counts[r.status] += 1
    
    not_marked = total - records.count()

    return Response({
        'date': str(target_date),
        'total': total,
        'not_marked': not_marked,
        **counts,
        'attendance_pct': round(((counts['present'] + counts['late']) / total * 100) if total else 0, 1),
    })


def _staff_scope_qs(user, org, campus_id=None):
    """Same staff-visibility rule as staff_attendance_list/summary, minus the
    org-admin default-to-principals-only narrowing — the dashboard widget and
    a staff member's own profile calendar need the full staff roster, not
    just principals, when no campus is picked."""
    staff_qs = User.objects.filter(organization=org, is_active=True).exclude(role__in=['student', 'superadmin', 'donor'])
    if user.is_principal():
        staff_qs = staff_qs.filter(campus=user.campus).exclude(role='org_admin')
    elif user.is_coordinator():
        staff_qs = staff_qs.filter(campus=user.campus, role='teacher')
    elif campus_id:
        staff_qs = staff_qs.filter(campus_id=campus_id)
    return staff_qs


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def staff_daily_attendance_stats(request):
    """Per-day present/absent counts over a trailing window, for the Staff
    Attendance dashboard widget. `?days=7` (weekly) / `?days=30` (monthly) /
    `?days=90|180` (3/6 months — the frontend buckets these into months,
    mirroring how the student WeeklyAttendanceChart handles the same ranges).
    `?campus_id=<id>` narrows to one campus.
    """
    from django.db.models import Count

    user = request.user
    org = getattr(user, 'organization', None)

    try:
        days = int(request.query_params.get('days', 7))
    except (TypeError, ValueError):
        days = 7
    days = max(1, min(days, 366))
    campus_id = request.query_params.get('campus_id')
    window_start = date.today() - timedelta(days=days - 1)

    staff_qs = _staff_scope_qs(user, org, campus_id)

    rows = (
        StaffAttendance.objects.filter(organization=org, date__gte=window_start, user__in=staff_qs)
        .values('date')
        .annotate(
            present=Count('id', filter=Q(status__in=['present', 'late'])),
            absent=Count('id', filter=Q(status='absent')),
        )
        .order_by('date')
    )

    result = []
    for r in rows:
        d = r.get('date')
        if not d:
            continue
        result.append({
            'day': d.strftime('%a'),
            'date': str(d),
            'present': r.get('present') or 0,
            'absent': r.get('absent') or 0,
        })
    return Response(result)


def _parse_date_range(request, default_days=7):
    """Shared start_date/end_date parsing for the staff history/export views."""
    today = date.today()
    end_str = request.query_params.get('end_date')
    start_str = request.query_params.get('start_date')
    try:
        end = datetime.strptime(end_str, '%Y-%m-%d').date() if end_str else today
    except ValueError:
        raise ValueError('Invalid end_date. Use YYYY-MM-DD.')
    try:
        start = datetime.strptime(start_str, '%Y-%m-%d').date() if start_str else end - timedelta(days=default_days - 1)
    except ValueError:
        raise ValueError('Invalid start_date. Use YYYY-MM-DD.')
    if start > end:
        raise ValueError('start_date must be before end_date.')
    if (end - start).days > 366:
        raise ValueError('Date range too large (max 366 days).')
    return start, end


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def staff_attendance_history(request):
    """
    GET /api/attendance/staff/history/?start_date=&end_date=&campus_id=

    One status per staff member per day, for the Weekly/Monthly grid view on
    the Staff Attendance page — mirrors the shape the student class-history
    grid consumes (a flat day list, each carrying that day's per-person
    statuses), just keyed by user instead of by classroom/student.
    """
    user = request.user
    org = getattr(user, 'organization', None)
    campus_id = request.query_params.get('campus_id')

    try:
        start, end = _parse_date_range(request, default_days=7)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    staff_qs = _staff_scope_qs(user, org, campus_id).select_related('campus').order_by('first_name', 'last_name')

    staff_list = [{
        'id': s.id,
        'name': s.get_full_name(),
        'employee_code': s.employee_code,
        'role': s.get_role_display(),
        'campus_name': s.campus.campus_name if s.campus else None,
    } for s in staff_qs]

    records = StaffAttendance.objects.filter(
        organization=org, date__gte=start, date__lte=end, user__in=staff_qs
    ).values('date', 'user_id', 'status')

    by_day: dict = {}
    for r in records:
        d = str(r['date'])
        by_day.setdefault(d, {})[str(r['user_id'])] = r['status']

    days = []
    for d in calendar_utils.date_range(start, end):
        d_str = str(d)
        days.append({'date': d_str, 'records': by_day.get(d_str, {})})

    return Response({
        'start_date': str(start),
        'end_date': str(end),
        'staff': staff_list,
        'days': days,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def staff_export_csv(request):
    """
    GET /api/attendance/staff/export-csv/?start_date=&end_date=&campus_id=

    Same matrix layout as export_attendance_csv (staff down the side, dates
    across the top, one status letter per cell) — feeds the Export button on
    the dashboard's Staff Attendance widget.
    """
    import csv

    user = request.user
    org = getattr(user, 'organization', None)
    campus_id = request.query_params.get('campus_id')

    try:
        start, end = _parse_date_range(request, default_days=7)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    staff_qs = _staff_scope_qs(user, org, campus_id).select_related('campus').order_by('campus__campus_name', 'first_name', 'last_name')

    records = StaffAttendance.objects.filter(
        organization=org, date__gte=start, date__lte=end, user__in=staff_qs
    ).values('date', 'user_id', 'status')
    by_user_day: dict = {}
    for r in records:
        by_user_day.setdefault(r['user_id'], {})[str(r['date'])] = r['status']

    dates = [str(d) for d in calendar_utils.date_range(start, end)]
    letters = {'present': 'P', 'absent': 'A', 'late': 'L', 'leave': 'Lv', 'half_day': 'HD'}

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="staff_attendance_{start}_to_{end}.csv"'
    writer = csv.writer(response)

    writer.writerow(['Date Range', f'{start} to {end}'])
    writer.writerow(['Legend', 'P=Present  A=Absent  L=Late  Lv=Leave  HD=Half Day  (blank)=not marked'])
    writer.writerow([])
    writer.writerow(['Staff', 'Employee Code', 'Role', 'Campus'] + [d[5:] for d in dates] + ['%'])

    for s in staff_qs:
        day_status = by_user_day.get(s.id, {})
        tally = {}
        for d in dates:
            st = day_status.get(d)
            if st:
                tally[st] = tally.get(st, 0) + 1
        eligible = sum(tally.values()) - tally.get('leave', 0)
        pct = round((tally.get('present', 0) + tally.get('late', 0)) / eligible * 100, 1) if eligible > 0 else 0
        row = [s.get_full_name(), s.employee_code, s.get_role_display(), s.campus.campus_name if s.campus else '']
        row += [letters.get(day_status.get(d), '') for d in dates]
        row.append(f'{pct}%')
        writer.writerow(row)

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employee_timings_list(request):
    """Return all staff members with their timing config."""
    from users.models import User
    user = request.user
    org = getattr(user, 'organization', None)

    campus_id = request.query_params.get('campus_id')
    qs = User.objects.filter(organization=org, is_active=True).exclude(role__in=['student', 'superadmin', 'donor'])
    
    if user.is_principal():
        qs = qs.filter(campus=user.campus).exclude(role='org_admin')
    elif user.is_coordinator():
        qs = qs.filter(campus=user.campus, role='teacher')
    elif user.is_org_admin_role():
        if campus_id:
            qs = qs.filter(campus_id=campus_id)
        else:
            qs = qs.filter(role='principal')

    result = []
    for u in qs.select_related('shift_timing', 'campus'):
        try:
            timing = u.shift_timing
            ci = str(timing.check_in_time)[:5] if timing.check_in_time else None
            co = str(timing.check_out_time)[:5] if timing.check_out_time else None
            grace = timing.grace_minutes
            tid = timing.id 
        except:
            ci, co, grace, tid = None, None, 10, None

        result.append({
            'user_id': u.id,
            'full_name': u.get_full_name(),
            'employee_code': u.employee_code,
            'campus': u.campus.campus_name if u.campus else "",
            'timing_id': tid,
            'check_in_time': ci,
            'check_out_time': co,
            'grace_minutes': grace,
        })
    return Response(result)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def employee_timings_bulk_save(request):
    """Bulk upsert employee timings."""
    user = request.user
    org = getattr(user, 'organization', None)

    if not (user.is_superuser or getattr(user, 'is_org_admin', False) or
            (hasattr(user, 'is_principal') and user.is_principal())):
        return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

    items = request.data if isinstance(request.data, list) else request.data.get('timings', [])
    saved = []
    for item in items:
        # Front-end might send teacher_id (legacy) or user_id
        target_id = item.get('user_id') or item.get('teacher_id')
        ci = item.get('check_in_time')
        co = item.get('check_out_time')
        grace = item.get('grace_minutes', 10)
        
        if not target_id or not ci or not co:
            continue
            
        try:
            target_user = User.objects.get(id=target_id, organization=org)
            obj, _ = EmployeeShiftTiming.objects.update_or_create(
                user=target_user,
                defaults={
                    'check_in_time': ci, 
                    'check_out_time': co, 
                    'grace_minutes': grace,
                    'organization': org, 
                    'is_active': True
                }
            )
            saved.append(obj.id)
        except Exception as e:
            print(f"Error saving timing for user {target_id}: {e}")
            continue
            
    return Response({'saved': len(saved)})


import csv

@api_view(['GET'])
# An export is the easiest thing here to carry out of the building, so it gets
# the same gate as viewing: the `view_attendance` toggle, not just a login.
@permission_classes([IsAuthenticated, HasAttendanceViewPermission])
def export_attendance_csv(request):
    """
    Export attendance records to CSV for a specific campus and date range.
    Principals can only export for their own campus.
    """
    from campus.models import Campus
    from attendance.models import StudentAttendance, Attendance
    
    from attendance.services.export_scope import (
        ExportScopeError, date_list as build_date_list, resolve_export_scope,
    )
    from classes.models import ClassRoom
    from students.models import Student
    from attendance.models import Attendance

    campus_id = request.GET.get('campus')

    # One answer to "what may this user see", shared with the review endpoint.
    # The old block pinned a Principal to their campus but left a Coordinator
    # able to export any campus — or, with no campus param, the whole org.
    try:
        classrooms, s_date, e_date, scope = resolve_export_scope(request)
    except ExportScopeError as exc:
        return Response({'error': exc.message}, status=exc.status_code)

    date_list = build_date_list(s_date, e_date)

    classrooms = classrooms.select_related('class_teacher', 'grade', 'grade__level').order_by('class_teacher__full_name', 'grade__name', 'section')

    # Fetch campus object for header info
    campus_obj = None
    if campus_id:
        try:
            from campus.models import Campus
            campus_obj = Campus.objects.get(id=campus_id)
        except:
            pass

    # Matrix layout — students down the side, dates across the top, a status
    # letter per cell — reusing the same builder the review page renders, so the
    # file and the screen agree. The old one-row-per-student-per-day format ran
    # to hundreds of "NOT MARKED" lines that nobody read.
    from attendance.services.export_matrix import (
        STATUS_LETTER, cell_text, iter_classroom_matrices, teacher_name,
    )

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{timezone.now().date()}.csv"'
    writer = csv.writer(response)

    writer.writerow(['Campus', campus_obj.campus_name if campus_obj else 'All Campuses'])
    writer.writerow(['Date Range', f"{s_date} to {e_date}"])
    writer.writerow(['Legend', 'P=Present  A=Absent  L=Late  Lv=Leave  Ex=Excused  H=Holiday  (blank)=weekend/not marked'])
    writer.writerow([])

    any_class = False
    for classroom, matrix in iter_classroom_matrices(classrooms, s_date, e_date, scope):
        any_class = True
        writer.writerow([f'Class: {classroom}'])
        writer.writerow([f'Teacher: {teacher_name(classroom)}'])
        writer.writerow([f"Working days: {matrix['working_days']}"])

        # Two header rows so it reads like the screen: day name on top, the
        # day-of-month below. MM-DD in one line was the old, denser look.
        writer.writerow(['Student', 'GR No'] + [d['day'] for d in matrix['dates']] + ['%'])
        writer.writerow(['', ''] + [d['date'][8:10] for d in matrix['dates']] + [''])

        for student in matrix['students']:
            row = [student['student_name'], student['gr_no']]
            row += [cell_text(matrix['dates'], student, d['date']) for d in matrix['dates']]
            row.append(f"{student['attendance_pct']}%")
            writer.writerow(row)

        writer.writerow([])  # blank line between classes

    if not any_class:
        writer.writerow(['No classes in scope for this range.'])

    return response


@api_view(['GET'])
# An export is the easiest thing here to carry out of the building, so it gets
# the same gate as viewing: the `view_attendance` toggle, not just a login.
@permission_classes([IsAuthenticated, HasAttendanceViewPermission])
def export_attendance_excel(request):
    """
    Export attendance records to professional Excel format (.xlsx).
    Supports bulk exports for multiple grades.
    """
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    from attendance.services.export_scope import ExportScopeError, resolve_export_scope
    from attendance.services.export_matrix import cell_text, iter_classroom_matrices, teacher_name

    # Same scope resolution as the CSV export and the review endpoint.
    try:
        classrooms, s_date, e_date, scope = resolve_export_scope(request)
    except ExportScopeError as exc:
        return Response({'error': exc.message}, status=exc.status_code)

    # Cell fills matching the on-screen roll, keyed by status letter.
    CELL_FILL = {
        'P': 'C6EFCE',   # green
        'A': 'FFC7CE',   # red
        'L': 'FFE0B2',   # orange
        'Lv': 'E1BEE7',  # purple
        'Ex': 'B2DFDB',  # teal
        'H': 'CFD8DC',   # blue-grey (holiday)
    }
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sheet_name(classroom, used):
        """Excel sheet titles: <= 31 chars, no []:*?/\\, and unique."""
        base = re.sub(r'[\[\]:*?/\\]', '', f"{classroom.grade.name} {classroom.section}")[:28]
        name, i = base or "Class", 1
        while name in used:
            name = f"{base[:26]}_{i}"
            i += 1
        used.add(name)
        return name

    wb = Workbook()
    wb.remove(wb.active)  # start clean; a sheet per class is added below
    used_names = set()
    any_class = False

    for classroom, matrix in iter_classroom_matrices(classrooms, s_date, e_date, scope):
        any_class = True
        ws = wb.create_sheet(sheet_name(classroom, used_names))

        # Title block.
        ws['A1'] = f"Class: {classroom}"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = f"Teacher: {teacher_name(classroom)}"
        ws['A3'] = f"{s_date} to {e_date}  ·  {matrix['working_days']} working days"

        # Header row: Student | GR No | dates… | %.
        header_row = 5
        headers = ['Student', 'GR No'] + [f"{d['day']}\n{d['date'][8:10]}" for d in matrix['dates']] + ['%']
        for col, title in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Student rows.
        row = header_row + 1
        for student in matrix['students']:
            ws.cell(row=row, column=1, value=student['student_name']).border = border
            ws.cell(row=row, column=2, value=student['gr_no']).border = border
            for i, d in enumerate(matrix['dates']):
                letter = cell_text(matrix['dates'], student, d['date'])
                cell = ws.cell(row=row, column=3 + i, value=letter)
                cell.alignment = center
                cell.border = border
                if letter in CELL_FILL:
                    cell.fill = PatternFill(
                        start_color=CELL_FILL[letter], end_color=CELL_FILL[letter], fill_type="solid",
                    )
            pct_cell = ws.cell(row=row, column=3 + len(matrix['dates']), value=f"{student['attendance_pct']}%")
            pct_cell.alignment = center
            pct_cell.border = border
            row += 1

        # Widths: name wide, date columns narrow, freeze header + name column.
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 16
        for i in range(len(matrix['dates'])):
            ws.column_dimensions[ws.cell(row=header_row, column=3 + i).column_letter].width = 5
        ws.freeze_panes = ws.cell(row=header_row + 1, column=3)

    if not any_class:
        ws = wb.create_sheet("Attendance")
        ws['A1'] = "No classes in scope for this range."

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{timezone.now().date()}.xlsx"'
    wb.save(response)
    return response


# ─── Audit Log: Extended List + Recovery ─────────────────────────────────────

class AuditLogListView(APIView):
    """
    GET /api/attendance/audit-logs/extended/
    Full paginated audit log for org-admin, superadmin, principal.
    Supports filters: action, feature, entity_type, start_date, end_date, search
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from .models import AuditLog
        from .serializers import DeleteLogSerializer
        from django.db.models import Q as Qobj

        user = request.user
        allowed_roles = ['superadmin', 'admin', 'org_admin', 'principal']
        if not (user.is_superuser or getattr(user, 'role', '') in allowed_roles):
            return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        qs = AuditLog.objects.select_related('user').order_by('-timestamp')

        # Filters
        action_f = request.query_params.get('action')
        feature_f = request.query_params.get('feature')
        entity_type_f = request.query_params.get('entity_type')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        search = request.query_params.get('search', '').strip()

        if action_f:
            qs = qs.filter(action=action_f)
        if feature_f:
            qs = qs.filter(feature=feature_f)
        if entity_type_f:
            qs = qs.filter(entity_type__icontains=entity_type_f)
        if start_date:
            qs = qs.filter(timestamp__date__gte=start_date)
        if end_date:
            qs = qs.filter(timestamp__date__lte=end_date)
        if search:
            qs = qs.filter(
                Qobj(reason__icontains=search) |
                Qobj(changes__icontains=search) |
                Qobj(entity_type__icontains=search)
            )

        # Pagination
        page_size = min(int(request.query_params.get('page_size', 50)), 200)
        page = max(int(request.query_params.get('page', 1)), 1)
        total = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        items = qs[start:end]

        serializer = DeleteLogSerializer(items, many=True)
        return Response({
            'results': serializer.data,
            'count': len(serializer.data),
            'total': total,
            'page': page,
            'page_size': page_size,
            'num_pages': max(1, (total + page_size - 1) // page_size),
        })


class RecoverEntityView(APIView):
    """
    POST /api/attendance/audit-logs/<id>/recover/
    Restores the soft-deleted entity referenced by the AuditLog entry.
    Only org-admin and superadmin can recover.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        from .models import AuditLog
        from django.db.models import Q as Qobj

        user = request.user
        allowed_roles = ['superadmin', 'admin', 'org_admin']
        if not (user.is_superuser or getattr(user, 'role', '') in allowed_roles):
            return Response({'error': 'Only org-admin or superadmin can restore records.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            log = AuditLog.objects.with_deleted().get(pk=pk) if hasattr(AuditLog.objects, 'with_deleted') else AuditLog.objects.get(pk=pk)
        except AuditLog.DoesNotExist:
            return Response({'error': 'Audit log entry not found.'}, status=status.HTTP_404_NOT_FOUND)

        if log.action != 'delete':
            return Response({'error': 'Only delete log entries can be recovered.'}, status=status.HTTP_400_BAD_REQUEST)

        if log.is_recovered:
            return Response({'error': 'This record has already been restored.'}, status=status.HTTP_400_BAD_REQUEST)

        entity_id = log.entity_id
        feature = log.feature
        restored = False
        entity_name = f"{log.entity_type} (ID: {entity_id})"

        try:
            if feature == 'student':
                from students.models import Student
                updated = Student.objects.with_deleted().filter(id=entity_id, is_deleted=True).update(
                    is_deleted=False, deleted_at=None
                )
                restored = updated > 0
                try:
                    entity_name = Student.objects.with_deleted().get(id=entity_id).name
                except Exception:
                    pass

            elif feature == 'teacher':
                from teachers.models import Teacher
                updated = Teacher.objects.with_deleted().filter(id=entity_id, is_deleted=True).update(
                    is_deleted=False, deleted_at=None
                )
                restored = updated > 0
                try:
                    entity_name = Teacher.objects.with_deleted().get(id=entity_id).full_name
                except Exception:
                    pass

            elif feature == 'result':
                from result.models import Result
                updated = Result.objects.with_deleted().filter(id=entity_id, is_deleted=True).update(
                    is_deleted=False, deleted_at=None
                )
                restored = updated > 0

            elif feature == 'attendance':
                updated = Attendance.objects.with_deleted().filter(id=entity_id, is_deleted=True).update(
                    is_deleted=False, deleted_at=None
                ) if hasattr(Attendance.objects, 'with_deleted') else 0
                restored = updated > 0

            else:
                return Response({'error': f'Recovery not supported for feature: {feature}'}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({'error': f'Recovery failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        if not restored:
            return Response({'error': 'Record not found or already active.'}, status=status.HTTP_404_NOT_FOUND)

        # Mark original log as recovered
        AuditLog.objects.filter(pk=pk).update(is_recovered=True)

        # Create restore audit log entry
        AuditLog.objects.create(
            feature=feature,
            action='restore',
            entity_type=log.entity_type,
            entity_id=entity_id,
            organization=log.organization,
            user=user,
            ip_address=request.META.get('REMOTE_ADDR'),
            changes={'restored_from_log_id': log.id, 'entity_name': entity_name},
            reason=f'{log.entity_type} (ID: {entity_id}) restored by {user.get_full_name() or user.username}'
        )

        return Response({
            'success': True,
            'message': f'"{entity_name}" has been restored successfully.',
            'entity_id': entity_id,
            'feature': feature,
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_campus_attendance_stats(request):
    """Attendance % per campus over a trailing window (default 30 days).
    `?days=7` for weekly. Used by the org-admin 'Campus Comparison' table."""
    from django.db.models import Count
    try:
        days = int(request.query_params.get('days', 30))
    except (TypeError, ValueError):
        days = 30
    days = max(1, min(days, 365))
    window_start = timezone.now().date() - timedelta(days=days)
    user = request.user
    qs = StudentAttendance.objects.filter(attendance__date__gte=window_start)
    org_id = getattr(user, 'organization_id', None)
    is_super = hasattr(user, 'is_superadmin') and user.is_superadmin()
    if org_id and not is_super:
        qs = qs.filter(attendance__classroom__grade__level__campus__organization_id=org_id)
    rows = (
        qs.values('attendance__classroom__grade__level__campus__campus_name')
          .annotate(
              total=Count('id', filter=Q(status__in=['present', 'absent', 'late'])),
              present=Count('id', filter=Q(status__in=['present', 'late'])),
          )
    )
    result = []
    for r in rows:
        name = r.get('attendance__classroom__grade__level__campus__campus_name')
        if not name:
            continue
        total = r.get('total') or 0
        present = r.get('present') or 0
        pct = round((present / total) * 100, 1) if total else 0.0
        result.append({'campus': name, 'percentage': pct, 'present': present, 'total': total})
    result.sort(key=lambda x: x['campus'])
    return Response(result)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_daily_attendance_stats(request):
    """Per-day present/absent counts over a trailing window, optionally scoped to
    one campus. Feeds the dashboard's Weekly Attendance chart.
    `?days=7` (weekly), `?campus=<id>` (INTRA drill-down)."""
    from django.db.models import Count
    try:
        days = int(request.query_params.get('days', 7))
    except (TypeError, ValueError):
        days = 7
    days = max(1, min(days, 366))
    campus = request.query_params.get('campus')
    window_start = timezone.now().date() - timedelta(days=days - 1)
    user = request.user
    qs = StudentAttendance.objects.filter(attendance__date__gte=window_start)
    org_id = getattr(user, 'organization_id', None)
    is_super = hasattr(user, 'is_superadmin') and user.is_superadmin()
    if org_id and not is_super:
        qs = qs.filter(attendance__classroom__grade__level__campus__organization_id=org_id)
    if campus:
        qs = qs.filter(attendance__classroom__grade__level__campus_id=campus)
    rows = (
        qs.values('attendance__date')
          .annotate(
              present=Count('id', filter=Q(status__in=['present', 'late'])),
              absent=Count('id', filter=Q(status='absent')),
          )
          .order_by('attendance__date')
    )
    result = []
    for r in rows:
        d = r.get('attendance__date')
        if not d:
            continue
        result.append({
            'day': d.strftime('%a'),
            'date': str(d),
            'present': r.get('present') or 0,
            'absent': r.get('absent') or 0,
        })
    return Response(result)
